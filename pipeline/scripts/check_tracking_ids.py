#!/usr/bin/env python3
"""Say which of a list of tracking IDs the source exports actually contain.

Tracking IDs arrive by hand -- in a mail, on a slide, pasted out of a planning
sheet -- and the question asked of them is always whether the activities behind
them exist. Searching the CSVs answers that for three IDs and stops working at
thirty, and it answers only half the question: an empty search says "not in
this file", not whether the activity was never created or whether the channel
suffix is three letters wrong. Those two answers lead somewhere different.

A match is exact on `tracking_id`, trimmed and upper-cased on both sides.
Nothing else counts as found. Every ID that does not match is then put through
a ladder of near-miss searches, and the first hit is reported as a hint beside
it -- never as a verdict.

The list is read as a workbook or as one ID per line, chosen by the extension.
The workbook is the easier file to keep by hand, and it can carry columns of
its own -- a campaign, a note, whoever asked -- which travel through to the
result file untouched. Every run writes that result: the answer is a file
someone sends on, not a flag to remember.

Usage (from the repo root, or just double-click trackids.cmd):
    python -m pipeline.scripts.check_tracking_ids --ids ids.xlsx
    python -m pipeline.scripts.check_tracking_ids --ids ids.txt --all
    python -m pipeline.scripts.check_tracking_ids --ids ids.xlsx --sheet "Q4"
    python -m pipeline.scripts.check_tracking_ids --ids ids.xlsx --out result.csv
    python -m pipeline.scripts.check_tracking_ids --ids ids.txt --input "C:\\path\\to\\Input"

Exit code 0 only when every listed ID was found.
"""

from __future__ import annotations

import argparse
import csv as csv_module
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# Where a run puts its workbook when the caller names no file. Same folder and
# same naming as the calendar report, which is already in .gitignore.
REPORTS_DIR = _REPO_ROOT / "pipeline" / "output" / "reports"

# Where a run looks for a list when the caller names no file: beside the
# launcher, which is where geb-members.xlsx already lives. Both names are in
# .gitignore -- a list carries campaign names and notes and is a working file,
# not repository content.
LIST_DIR = _REPO_ROOT

from pipeline.scripts.process_cplan import (  # noqa: E402
    find_input_dir,
    find_input_files,
    log,
    print_banner,
    print_kv,
    print_table,
    read_csv_auto,
    transform,
)


class IdListError(ValueError):
    """The list exists but cannot be read as one.

    Raised rather than falling back to an empty list: a run that searches
    nothing reports every ID as missing, which reads exactly like a real answer.
    """


@dataclass(frozen=True)
class IdList:
    """The IDs a list names, and whatever else its rows carried.

    `extras` is empty for a text list and holds the workbook's other columns
    otherwise, keyed on the normalised ID. They are carried rather than
    dropped so the result file can go back to whoever sent the list with their
    own columns still beside the answer.
    """

    listed: list[str]
    counts: Counter
    extras: dict[str, dict[str, str]]
    extra_columns: tuple[str, ...]


def normalise(value: str) -> str:
    """The one definition of "the same ID": trimmed, upper-cased."""
    return str(value).strip().upper()


# What the ID column may be called. `tacking id` is the export's own
# long-standing typo, which `transform()` already folds into `tracking_id` --
# a header pasted out of the export carries it, and refusing that header would
# make the export's mistake the operator's problem.
ID_HEADERS = ("tracking id", "tacking id")

# Searched in this order, so a folder holding one of them needs no flag. The
# workbook comes first only to make the search deterministic; holding both is
# an error rather than a precedence question -- see default_id_list.
DEFAULT_LIST_NAMES = ("ids.xlsx", "ids.txt")


def default_id_list(directory: Path) -> Path | None:
    """Which default list a directory holds, or None when it holds neither.

    Holding both is an error rather than a precedence rule. The two files would
    only both exist because one was converted from the other, and the moment
    they disagree, quietly reading the one the operator is not editing answers
    from a list nobody checked -- which is the silent-wrong-answer failure this
    whole check exists to prevent.
    """
    found = [directory / name for name in DEFAULT_LIST_NAMES if (directory / name).is_file()]
    if len(found) > 1:
        raise IdListError(
            f"{directory} holds both {DEFAULT_LIST_NAMES[0]} and {DEFAULT_LIST_NAMES[1]}; "
            f"keep one, or name the one to use with --ids"
        )
    return found[0] if found else None


def read_id_list(path: Path, sheet: str | None = None) -> IdList:
    """The IDs the file lists, in first-seen order, once each -- and how often.

    The extension picks the reader: `.xlsx` is read as a workbook, anything
    else as one ID per line. A repeat is not an error and not a second row: it
    is counted, and the count is what the report names.
    """
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx_list(path, sheet)
    if sheet is not None:
        raise IdListError(f"{path.name}: a sheet can only be chosen in an .xlsx list")
    return _read_text_list(path)


def _read_text_list(path: Path) -> IdList:
    """One ID per line.

    Blank lines and lines whose first non-space character is `#` are dropped,
    so a list can carry its own headings.
    """
    listed: list[str] = []
    counts: Counter = Counter()
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        key = normalise(line)
        if key not in counts:
            listed.append(line)
        counts[key] += 1
    return IdList(listed=listed, counts=counts, extras={}, extra_columns=())


def _require_openpyxl():
    """openpyxl, or a message naming the one command that fixes its absence.

    Imported here rather than at module scope so the text path -- and every
    machine that only ever uses it -- keeps working without the package.
    """
    try:
        import openpyxl
    except ImportError as error:
        raise IdListError(
            "reading an .xlsx list needs openpyxl: python -m pip install openpyxl"
        ) from error
    return openpyxl


def _text(value) -> str:
    """A cell as the string a person would have typed into it.

    Excel decides on its own that some cells are numbers or dates; whatever it
    hands back has to compare against the source data as text.
    """
    if value is None:
        return ""
    return str(value).strip()


def _read_xlsx_list(path: Path, sheet: str | None) -> IdList:
    """The same, from one sheet of a workbook.

    The ID column is found by its header, so the operator may reorder columns
    and add their own. Everything else on the row is carried along.
    """
    import zipfile

    openpyxl = _require_openpyxl()
    from openpyxl.utils.exceptions import InvalidFileException

    workbook = None
    try:
        # data_only: a formula cell must yield what Excel last displayed, not
        # its source text. read_only also keeps openpyxl from materialising the
        # blank rows Excel leaves behind below an edited-down list.
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet is None:
            worksheet = workbook.worksheets[0]
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            raise IdListError(
                f"{path.name}: no sheet named {sheet!r} "
                f"(it has: {', '.join(workbook.sheetnames)})"
            )

        rows = worksheet.iter_rows(values_only=True)
        headers = [_text(value) for value in (next(rows, None) or ())]
        id_column = next(
            (index for index, name in enumerate(headers) if name.lower() in ID_HEADERS),
            None,
        )
        if id_column is None:
            raise IdListError(
                f"{path.name}: no 'Tracking ID' column "
                f"(found: {', '.join(name for name in headers if name) or 'nothing'})"
            )
        extra_columns = tuple(
            name for index, name in enumerate(headers) if index != id_column and name
        )

        listed: list[str] = []
        counts: Counter = Counter()
        extras: dict[str, dict[str, str]] = {}
        for offset, raw in enumerate(rows, start=2):
            cells = [_text(value) for value in raw]
            # Excel hands back every row the sheet has ever held, so a list
            # edited down from a longer one arrives with blank rows below the
            # data. Skipping them is what makes the format usable by hand --
            # but only when the whole row is blank. A note with no ID beside it
            # is an ID someone meant to fill in, and reporting nothing about it
            # is the one outcome that helps no one.
            if not any(cells):
                continue
            value = cells[id_column] if id_column < len(cells) else ""
            if not value:
                raise IdListError(f"{path.name}: row {offset} carries no tracking ID")
            if value.startswith("#"):
                continue
            key = normalise(value)
            if key not in counts:
                listed.append(value)
                # First row wins, the same rule the order follows, so the two
                # cannot disagree about which row an ID came from.
                extras[key] = {
                    name: cells[index] if index < len(cells) else ""
                    for index, name in enumerate(headers)
                    if index != id_column and name
                }
            counts[key] += 1
        return IdList(
            listed=listed, counts=counts, extras=extras, extra_columns=extra_columns
        )
    except IdListError:
        raise
    except (OSError, zipfile.BadZipFile, InvalidFileException,
            ValueError, KeyError, TypeError) as error:
        # A CSV saved and renamed to .xlsx is the obvious wrong move; an .xls
        # renamed likewise is the other one. They surface as BadZipFile (which
        # descends straight from Exception, so it needs naming) and as
        # InvalidFileException respectively. Name the file either way -- a
        # traceback here says nothing about which file to go fix.
        raise IdListError(f"{path.name}: {error}") from error
    finally:
        if workbook is not None:
            workbook.close()


# The exports that carry a `tracking_id`, paired with the source type
# `transform()` reads them as, in the order a duplicate is resolved: an ID that
# is in both a live export and an archive is answered by the live row.
#
# The pack, channel and cluster exports are deliberately absent. They carry
# pack, channel and cluster identifiers, and searching them would let a pack ID
# report as a found activity.
ACTIVITY_SOURCES = (
    ("internal", "internal"),
    ("external", "external"),
    ("internal_archive", "internal"),
    ("external_archive", "external"),
)


@dataclass(frozen=True)
class Entry:
    """One activity the export carries, as much of it as the report shows."""

    tracking_id: str
    source: str
    sp_id: str
    activity_name: str


def _cell(row, column: str) -> str:
    """A column's value as a printable string, or "" where there is none."""
    if column not in row:
        return ""
    value = row[column]
    if value is None or (isinstance(value, float) and value != value):  # NaN
        return ""
    text = str(value).strip()
    return "" if text in ("nan", "None") else text


def build_index(files: dict[str, Path]) -> dict[str, Entry]:
    """Normalised tracking ID to the export row that carries it.

    Each file goes through the ETL's own `read_csv_auto()` and `transform()`.
    `transform()` is what turns the SharePoint-encoded headers into
    `tracking_id`, and what folds the export's long-standing `Tacking ID` typo
    variant into the same column -- reading the raw header would miss every row
    in whichever file carries the typo that week.
    """
    index: dict[str, Entry] = {}
    for key, source_type in ACTIVITY_SOURCES:
        path = files.get(key)
        if path is None:
            continue
        frame = transform(read_csv_auto(path), source_type)
        if "tracking_id" not in frame.columns:
            log(f"  {path.name} carries no tracking ID column")
            continue
        added = 0
        for _, row in frame.iterrows():
            tracking_id = normalise(_cell(row, "tracking_id"))
            if not tracking_id or tracking_id in index:
                continue
            index[tracking_id] = Entry(
                tracking_id=tracking_id,
                source=key,
                sp_id=_cell(row, "sp_id"),
                activity_name=_cell(row, "activity_name"),
            )
            added += 1
        log(f"  {key}: {added} tracking ID(s)")
    return index


# CLUSTER-PACKNUM-YYMMDD-ACTNUM-CHANNEL
PART_COUNT = 5


def _within_one_edit(left: str, right: str) -> bool:
    """True when one substitution, insertion or deletion turns one into the other."""
    if abs(len(left) - len(right)) > 1:
        return False
    if len(left) == len(right):
        differences = sum(1 for a, b in zip(left, right) if a != b)
        return differences == 1
    shorter, longer = (left, right) if len(left) < len(right) else (right, left)
    i = j = 0
    skipped = False
    while i < len(shorter) and j < len(longer):
        if shorter[i] == longer[j]:
            i += 1
            j += 1
            continue
        if skipped:
            return False
        skipped = True
        j += 1
    return True


@dataclass(frozen=True)
class Hint:
    """Why an ID may be missing, and the thing in the export to go look at.

    Two fields rather than one sentence, because the sentence is what the
    report has to print and a tracking ID is 32 characters wide. Joined into a
    column it is the ID that gets truncated -- and the ID is the whole point of
    the hint.
    """

    why: str = ""
    nearest: str = ""

    @property
    def text(self) -> str:
        """The one-string form, for the CSV where width costs nothing."""
        if self.why and self.nearest:
            return f"{self.why}: {self.nearest}"
        return self.why or self.nearest

    def __bool__(self) -> bool:
        return bool(self.why or self.nearest)


def find_hint(wanted: str, index: dict[str, Entry]) -> Hint:
    """Why this ID may be missing -- the first rung that hits, or an empty Hint.

    Never a verdict. The row still reads `missing`; this only says where to
    look, because "not found" and "found, spelled differently" lead somewhere
    completely different.
    """
    key = normalise(wanted)
    parts = key.split("-")
    why = ""

    if len(parts) == PART_COUNT:
        pack = f"{parts[0]}-{parts[1]}"
        activity_number = parts[3]

        # Rung 1: the same activity, published on another channel.
        for candidate in index:
            other = candidate.split("-")
            if len(other) != PART_COUNT:
                continue
            if f"{other[0]}-{other[1]}" == pack and other[3] == activity_number:
                return Hint(f"wrong channel, it is {other[4]}", candidate)

        # Rung 2: the pack exists, this activity within it does not.
        in_pack = sum(1 for candidate in index if candidate.startswith(f"{pack}-"))
        if in_pack:
            return Hint(f"pack exists ({in_pack}), this does not", pack)
    else:
        why = f"wrong shape ({len(parts)} parts, not {PART_COUNT})"

    # Rung 3: one character off. Still run for a malformed ID -- a note saying
    # only "wrong shape" leaves the typo it is a symptom of unfound.
    for candidate in index:
        if _within_one_edit(key, candidate):
            return Hint(why or "one character off", candidate)

    return Hint(why)


@dataclass(frozen=True)
class Result:
    """One listed ID, and what the export had to say about it."""

    listed: str
    entry: Entry | None
    hint: Hint
    times_listed: int
    extras: dict[str, str]

    @property
    def status(self) -> str:
        return "found" if self.entry else "missing"


def check(id_list: IdList, index: dict[str, Entry]) -> list[Result]:
    """Each listed ID against the index, in the order the list gave them."""
    results = []
    for value in id_list.listed:
        key = normalise(value)
        entry = index.get(key)
        results.append(
            Result(
                listed=value,
                entry=entry,
                hint=Hint() if entry else find_hint(key, index),
                times_listed=id_list.counts[key],
                extras=id_list.extras.get(key, {}),
            )
        )
    return results


def report(results: list[Result], show_all: bool) -> None:
    """The three numbers, then the rows that need doing something about.

    The found ones stay a count. The list is something the reader already has;
    printing it back sorted into two piles makes them read forty rows to find
    the three that matter.
    """
    missing = [r for r in results if r.entry is None]
    found = [r for r in results if r.entry is not None]

    print_kv([
        ("Searched", len(results)),
        ("Found", len(found)),
        ("Missing", len(missing)),
    ])
    print()

    repeated = [r for r in results if r.times_listed > 1]
    if repeated:
        log(f"{len(repeated)} ID(s) listed more than once; each was searched once")

    if missing:
        print_table(
            "Missing",
            ["Tracking ID", "Why it may be missing", "Nearest in export"],
            [
                (
                    r.listed,
                    r.hint.why or ("" if r.hint.nearest else "nothing close"),
                    r.hint.nearest,
                )
                for r in missing
            ],
            col_widths=[34, 32, 34],
        )

    if show_all and found:
        print_table(
            "Found",
            ["Tracking ID", "Source", "SP ID", "Activity"],
            [(r.listed, r.entry.source, r.entry.sp_id, r.entry.activity_name) for r in found],
            col_widths=[36, 20, 9, 42],
        )

    if missing:
        log(f"{len(missing)} of {len(results)} ID(s) are not in the export.")
    else:
        log(f"OK: all {len(results)} ID(s) are in the export.")
    print()


CSV_COLUMNS = ("id", "status", "source_file", "sp_id", "activity_name", "hint")

# What --out understands. The extension picks the writer, so a caller gets the
# format they named rather than the one this script would have preferred.
WRITERS = (".xlsx", ".csv")


def default_output_path() -> Path:
    """Where a run puts its workbook when the caller names no file."""
    stamp = datetime.now().strftime("%Y_%m_%d")
    return REPORTS_DIR / f"CPLAN_trackids_{stamp}.xlsx"


def _rows(results: list[Result], extra_columns: tuple[str, ...]) -> list[list[str]]:
    """Every row, found and missing alike -- a file is read by a spreadsheet.

    The list's own columns follow the fixed ones, in the order the list had
    them, so the file can go back to whoever sent it and still be recognisable.
    """
    rows = []
    for result in results:
        entry = result.entry
        rows.append([
            result.listed,
            result.status,
            entry.source if entry else "",
            entry.sp_id if entry else "",
            entry.activity_name if entry else "",
            result.hint.text,
            *(result.extras.get(column, "") for column in extra_columns),
        ])
    return rows


def write_result(path: Path, results: list[Result], extra_columns: tuple[str, ...]) -> None:
    """The result file, in the format its extension names."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        write_csv(path, results, extra_columns)
    else:
        write_xlsx(path, results, extra_columns)


def write_csv(path: Path, results: list[Result], extra_columns: tuple[str, ...] = ()) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv_module.writer(handle)
        writer.writerow([*CSV_COLUMNS, *extra_columns])
        writer.writerows(_rows(results, extra_columns))


def write_xlsx(path: Path, results: list[Result], extra_columns: tuple[str, ...] = ()) -> None:
    """The same rows as one sheet, frozen and filtered.

    Imported here rather than at module scope because `pipeline.report.style`
    imports openpyxl on sight: at the top of this file it would make openpyxl
    a hard requirement of the text-list path, which is the one path that has
    always run without it.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    from pipeline.report import style

    headers = [*CSV_COLUMNS, *extra_columns]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Result"
    row = style.write_header_row(sheet, 1, headers)
    style.write_data_rows(sheet, row, _rows(results, extra_columns))
    # Filtering down to the missing rows is the one thing this file is opened
    # for, and it has to cover the header row for Excel to name the columns.
    last = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last}{len(results) + 1}"
    style.finalize_sheet(sheet)
    workbook.save(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--ids",
        type=Path,
        default=None,
        help="the tracking IDs: an .xlsx with a 'Tracking ID' column, or a text "
             f"file with one per line. Default: {' or '.join(DEFAULT_LIST_NAMES)} "
             "beside the launcher",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="read the CSVs from this folder instead of the usual OneDrive/local discovery",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="read this sheet of an .xlsx list instead of the first one",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="also list the IDs that were found, not only the count",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help=f"write the full result here; the extension picks the format "
             f"({', '.join(WRITERS)}). Default: {REPORTS_DIR.name}/CPLAN_trackids_<date>.xlsx",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="the older spelling of --out, kept because notes and scripts carry it",
    )
    args = parser.parse_args(argv)

    print_banner("CPLAN tracking-ID check")

    out_path = args.out or args.csv or default_output_path()
    # Checked before anything is read: a mistyped extension is worth a second,
    # not the twenty this spends indexing four exports first.
    if out_path.suffix.lower() not in WRITERS:
        log(f"ERROR: cannot write {out_path.suffix or 'a file with no extension'} "
            f"-- name a file ending in {' or '.join(WRITERS)}")
        print()
        return 1

    try:
        ids_path = args.ids if args.ids is not None else default_id_list(LIST_DIR)
    except IdListError as error:
        log(f"ERROR: {error}")
        print()
        return 1

    if ids_path is None:
        log(f"ERROR: no ID list. Put an {DEFAULT_LIST_NAMES[0]} beside the launcher "
            f"(or an {DEFAULT_LIST_NAMES[1]}), or name one with --ids.")
        print()
        return 1

    if not ids_path.is_file():
        log(f"ERROR: no such ID list: {ids_path}")
        print()
        return 1

    try:
        id_list = read_id_list(ids_path, args.sheet)
    except IdListError as error:
        log(f"ERROR: {error}")
        print()
        return 1
    listed = id_list.listed
    if not listed:
        # The sheet is named when one was chosen: with several to pick from,
        # which one was read is the whole question behind an empty answer.
        where = ids_path.name + (f" sheet {args.sheet!r}" if args.sheet else "")
        log(f"ERROR: no tracking IDs in {where} -- every row was blank, commented or a header.")
        print()
        return 1
    print_kv([("ID list", str(ids_path)), ("IDs to search", len(listed))])
    print()

    if args.input is not None:
        if not args.input.is_dir():
            log(f"ERROR: not a folder: {args.input}")
            print()
            return 1
        input_dir = args.input
        log(f"Using input: {input_dir}")
    else:
        input_dir = find_input_dir()

    files = find_input_files(input_dir)
    if not any(key in files for key, _ in ACTIVITY_SOURCES):
        log("ERROR: no activity export found.")
        print_kv([("Input dir", str(input_dir))])
        print()
        return 1

    index = build_index(files)
    print()
    results = check(id_list, index)
    report(results, args.all)

    write_result(out_path, results, id_list.extra_columns)
    log(f"Result written to {out_path}")
    print()

    return 0 if all(r.entry for r in results) else 1


if __name__ == "__main__":
    sys.exit(main())
