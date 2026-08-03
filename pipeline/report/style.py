"""Palette and write primitives.

Every sheet is composed from these six writers plus finalize_sheet; that is the
only reason the workbook reads as one document rather than seven.

Colours come from the corporate design system and match what the studio's own
.xlsx export already writes -- white dominates, greys carry structure, the
accent appears only on the summary tab.
"""

import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WHITE = "FFFFFF"
BLACK = "000000"
GRAY_I = "CCCABC"     # borders
GRAY_IV = "7A7870"    # tab colour, sub-labels
GRAY_VI = "404040"    # header fill
PASTEL_I = "ECEBE4"   # section bands and total rows
ROW_ALT = "F8F7F2"    # zebra striping
BRONZE_I = "B98E2C"   # the summary tab only

NUM_FMT_INT = "#,##0"
NUM_FMT_PCT = "0.0%"
NUM_FMT_RATIO = "0.0"
NUM_FMT_DATE = "YYYY-MM-DD"

MIN_WIDTH = 10
MAX_WIDTH = 40

HEADER_FONT = Font(bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=GRAY_VI, end_color=GRAY_VI, fill_type="solid")
SECTION_FONT = Font(bold=True, color=GRAY_VI, size=11)
SECTION_FILL = PatternFill(start_color=PASTEL_I, end_color=PASTEL_I, fill_type="solid")
ALT_FILL = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
TOTAL_FILL = SECTION_FILL
TOTAL_FONT = Font(bold=True, color=BLACK, size=11)
LABEL_FONT = Font(bold=True, color=GRAY_VI, size=11)
SUB_FONT = Font(italic=True, color=GRAY_IV, size=10)
BODY_FONT = Font(size=11)
_SIDE = Side(style="thin", color=GRAY_I)
THIN_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


# openpyxl raises on control characters rather than escaping them, and the
# exception lands mid-write: the run dies and no workbook is produced at all.
# The ETL strips them at the source (process_cplan.strip_control_chars); this is
# the backstop for text the report builds itself, so a bad byte can at worst
# blank a character rather than lose the whole file.
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def safe(value):
    """A value openpyxl will accept. Non-strings pass through untouched."""
    if isinstance(value, str):
        return _ILLEGAL.sub("", value)
    return value


def write_header_row(ws, row, headers, col_start=1):
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + offset, value=safe(header))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    return row + 1


def write_data_rows(ws, row, rows, fmt_map=None, col_start=1):
    fmt_map = fmt_map or {}
    for index, values in enumerate(rows):
        fill = ALT_FILL if index % 2 == 1 else None
        for offset, value in enumerate(values):
            column = col_start + offset
            cell = ws.cell(row=row + index, column=column, value=safe(value))
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if column in fmt_map:
                cell.number_format = fmt_map[column]
            elif isinstance(value, bool):
                pass
            elif isinstance(value, float):
                cell.number_format = NUM_FMT_RATIO
            elif isinstance(value, int):
                cell.number_format = NUM_FMT_INT
    return row + len(rows)


def write_section_header(ws, row, title, span, col_start=1):
    for offset in range(span):
        cell = ws.cell(row=row, column=col_start + offset)
        cell.fill = SECTION_FILL
        if offset == 0:
            cell.value = title
            cell.font = SECTION_FONT
    return row + 1


def write_kpi_row(ws, row, label, value, fmt=NUM_FMT_INT, sub=False):
    """Label in column A, value in column B. `sub` indents and italicises."""
    label_cell = ws.cell(row=row, column=1, value=safe(label))
    label_cell.font = SUB_FONT if sub else LABEL_FONT
    label_cell.alignment = Alignment(indent=4 if sub else 1)
    value_cell = ws.cell(row=row, column=2, value=safe(value))
    value_cell.font = SUB_FONT if sub else BODY_FONT
    value_cell.alignment = Alignment(horizontal="right")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value_cell.number_format = fmt
    return row + 1


def write_formula(ws, row, col, formula, fmt=None, fill=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if bold:
        cell.font = TOTAL_FONT
    return cell


def note_missing(ws, message):
    """Graceful degradation: one honest cell instead of a traceback."""
    cell = ws.cell(row=1, column=1, value=safe(message))
    cell.font = SUB_FONT
    return 2


def auto_fit_columns(ws, min_width=MIN_WIDTH, max_width=MAX_WIDTH):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, min_width), max_width)


def finalize_sheet(ws, freeze="A2", widths=None):
    """Freeze, then set widths. Explicit widths win over the auto fit."""
    if freeze:
        ws.freeze_panes = freeze
    auto_fit_columns(ws)
    for letter, width in (widths or {}).items():
        ws.column_dimensions[letter].width = width
