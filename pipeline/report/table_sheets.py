"""The flat sheets: summary, data quality, audience, mix, activities, glossary.

The calendar lives in its own module; this file holds the six sheets that are
label/value lists and tables. They share the style primitives, so adding a block
is a handful of lines rather than a new layout.
"""

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from pipeline.report import metrics, regions, style
from pipeline.report.config import (
    AUDIENCE_BAND_ORDER,
    AUDIENCE_BANDS,
    BAND_UNKNOWN,
    LARGE_AUDIENCE_BANDS,
    SHORT_NOTICE_DAYS,
)
from pipeline.report.data import (
    COMPLETENESS_FIELDS_COMMON,
    COMPLETENESS_FIELDS_INTERNAL,
    EXCLUSION_ORDER,
)
from pipeline.report.derive import priority_rank, split_multi, split_people


def _share_label(total_row, value_row, text):
    """The share rides in the label so the value column stays a column of counts."""
    return f'=TEXT(IF(B${total_row}=0,0,B{value_row}/B${total_row}),"0%") & "  {text}"'


def _write_share_row(ws, row, total_row, text, count, sub=True):
    """A count with its share riding in the label.

    `sub` indents and italicises it as a member of the count above -- right for
    a breakdown, wrong for a headline figure that happens to be a share of the
    portfolio rather than of the row above it.
    """
    ws.cell(row=row, column=1, value=_share_label(total_row, row, text))
    style.write_kpi_row(ws, row, None, count, sub=sub)


def build_executive_summary(wb, scope, config):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = style.BRONZE_I
    frame = scope.frame
    row = 1

    row = style.write_section_header(ws, row, "REPORT", 2)
    for label, value in config.describe():
        row = style.write_kpi_row(ws, row, label, value)
    row = style.write_kpi_row(ws, row, "Weeks covered", len(scope.grid.weeks))
    # No "Source: <file>" rows. They named the operator's export files to an
    # audience that never sees that directory, and a filename carrying a date
    # or a person's initials invites a question the workbook cannot answer.
    # `report_calendar.py` logs them to the console instead, where the person
    # who chose the files is the one reading. Removed 2026-08-06.
    row = style.write_kpi_row(ws, row, "Rows read", scope.rows_read)
    for reason in EXCLUSION_ORDER:
        row = style.write_kpi_row(ws, row, f"Excluded: {reason}", scope.excluded[reason])
    row = style.write_kpi_row(ws, row, "Activities in scope", len(frame))
    row += 1

    row = style.write_section_header(ws, row, "VOLUME", 2)
    total_row = row
    row = style.write_kpi_row(ws, row, "Activities in scope", len(frame))
    for source_type in ("internal", "external"):
        count = int((frame.get("source_type") == source_type).sum()) if len(frame) else 0
        _write_share_row(ws, row, total_row, source_type.title(), count)
        row += 1
    for band in AUDIENCE_BAND_ORDER:
        count = int((frame.get("audience_band") == band).sum()) if len(frame) else 0
        _write_share_row(ws, row, total_row, band, count)
        row += 1
    row += 1

    stats = metrics.load_stats(scope)
    row = style.write_section_header(ws, row, "LOAD", 2)
    row = style.write_kpi_row(ws, row, "Median activities per week",
                              stats["median_per_week"], fmt=style.NUM_FMT_RATIO)
    row = style.write_kpi_row(ws, row, "Peak week", stats["peak_week_label"])
    row = style.write_kpi_row(ws, row, "Activities in the peak week", stats["peak_week_count"])
    row = style.write_kpi_row(ws, row, "Weeks with no activity", stats["zero_weeks"])
    row = style.write_kpi_row(ws, row, "Longest run of empty weeks", stats["longest_zero_run"])
    row = style.write_kpi_row(ws, row, "Share in the five busiest weeks",
                              stats["top5_share"], fmt=style.NUM_FMT_PCT)
    row += 1

    # Both figures carry their share in the label, the same way VOLUME does, so
    # column C is not one lone percentage next to two bare counts. No separate
    # "Audience band unknown" row: VOLUME above now lists every band including
    # Unknown, and printing the same number twice on one sheet reads as an error.
    row = style.write_section_header(ws, row, "LEADERSHIP & AUDIENCE", 2)
    executives = int(frame["has_executives"].sum()) if len(frame) else 0
    _write_share_row(ws, row, total_row, "With GEB/GEB-1 involvement", executives, sub=False)
    row += 1
    if scope.membership is not None:
        # Indented under the combined figure: it is a part of it, not a
        # second independent measure.
        geb = int((frame["executives_geb"] != "").sum()) if len(frame) else 0
        _write_share_row(ws, row, total_row, "With GEB involvement", geb, sub=True)
        row += 1
    large = int(frame["audience_band"].isin(LARGE_AUDIENCE_BANDS).sum()) if len(frame) else 0
    _write_share_row(ws, row, total_row, "Large audience (top two bands)", large,
                     sub=False)
    row += 2

    lead = metrics.lead_time_stats(frame) if len(frame) else {
        "counted": 0, "median_days": None, "short_notice": 0}
    row = style.write_section_header(ws, row, "PLANNING DISCIPLINE", 2)
    row = style.write_kpi_row(ws, row, "Median lead time (days)",
                              lead["median_days"] if lead["median_days"] is not None else "—")
    row = style.write_kpi_row(ws, row,
                              f"Planned at under {SHORT_NOTICE_DAYS} days' notice",
                              lead["short_notice"])
    row = style.write_kpi_row(ws, row, "Lead time not measurable",
                              len(frame) - lead["counted"])
    row += 1

    packs = metrics.pack_stats(frame) if len(frame) else {"without_pack": 0}
    row = style.write_section_header(ws, row, "DATA QUALITY", 2)
    median_completeness = int(frame["completeness"].median()) if len(frame) else 0
    row = style.write_kpi_row(ws, row, "Median planning completeness (%)", median_completeness)
    row = style.write_kpi_row(ws, row, "Without a pack link", packs["without_pack"])
    row += 1

    # The same signature the studio and the portal carry. This workbook travels
    # furthest -- it reaches people who open neither application -- and a mark
    # that reads differently here would not accumulate into recognition. A
    # label-only row, so it reads as a footer rather than as one more figure.
    style.write_kpi_row(ws, row, "Produced by ECC Measurement & Insights", None)

    style.finalize_sheet(ws, freeze="A2", widths={"A": 44, "B": 24, "C": 12})


def build_data_quality(wb, scope, config):
    """The pack problem, quantified: bucket sizes a planner can act on."""
    ws = wb.create_sheet("Data Quality")
    frame = scope.frame
    if frame.empty:
        style.note_missing(ws, "No activities in scope for the configured criteria")
        style.finalize_sheet(ws, freeze="A2")
        return

    row = style.write_section_header(ws, 1, "FIELD COMPLETENESS", 4)
    row = style.write_header_row(ws, row, ["Field", "Filled", "Missing", "% missing"])
    for name, filled, missing in metrics.field_completeness(scope):
        style.write_data_rows(ws, row, [[name, filled, missing]])
        style.write_formula(ws, row, 4, f"=IF(B{row}+C{row}=0,0,C{row}/(B{row}+C{row}))",
                            fmt=style.NUM_FMT_PCT)
        row += 1
    row += 1

    # Its own block, deliberately: the rates above are one per *reported* field
    # (metrics.REPORTED_FIELDS -- a wider list chosen for what a planner wants
    # to see missing), while this median is computed over the fields the entry
    # form requires, split by source type. Two correct numbers over two
    # different denominators sitting in one block read as a contradiction --
    # on the test fixture the block shows bod_geb 93% missing and then 100%
    # median completeness. Naming the denominator here is what makes them
    # reconcilable.
    row = style.write_section_header(ws, row, "PLANNING COMPLETENESS", 4)
    row = style.write_header_row(ws, row, ["Measure", "Value", "", ""])
    row = style.write_data_rows(ws, row, [
        ["Median planning completeness (%)", int(frame["completeness"].median())]])
    common = [name for name in COMPLETENESS_FIELDS_COMMON
              if name in scope.completeness_fields]
    internal_only = [name for name in COMPLETENESS_FIELDS_INTERNAL
                     if name in scope.completeness_fields
                     and name not in COMPLETENESS_FIELDS_COMMON]
    row = style.write_data_rows(ws, row, [
        ["Counted for every activity", ", ".join(common) or "—"],
        ["Counted for internal activities in addition", ", ".join(internal_only) or "—"],
        ["Excluded: not carried by this export",
         ", ".join(scope.skipped_completeness_fields) or "—"],
    ])
    row += 1

    packs = metrics.pack_stats(frame)
    row = style.write_section_header(ws, row, "PACK COVERAGE", 4)
    row = style.write_header_row(ws, row, ["Measure", "Count", "", ""])
    pack_rows = [
        ("Activities with a pack link", packs["with_pack"]),
        ("Activities without a pack link", packs["without_pack"]),
        ("Distinct packs", packs["packs"]),
        ("Packs holding exactly one activity", packs["singleton_packs"]),
        ("Packs holding 2 to 10", packs["small_packs"]),
        ("Packs holding 11 to 50", packs["medium_packs"]),
        ("Packs holding more than 50", packs["oversized_packs"]),
        ("Largest pack", packs["largest_pack"]),
    ]
    row = style.write_data_rows(ws, row, [[label, value] for label, value in pack_rows])
    row += 1

    row = style.write_section_header(ws, row, "RECORD ANOMALIES", 4)
    row = style.write_header_row(ws, row, ["Anomaly", "Count", "", ""])
    row = style.write_data_rows(ws, row, [[label, count] for label, count in
                                          metrics.anomalies(frame, scope.duplicates_removed)])
    row += 1

    row = _write_unmapped_regions(ws, row, frame) + 1

    if scope.membership is not None:
        row = style.write_section_header(ws, row, "GEB LIST", 4)
        row = style.write_header_row(ws, row, ["Measure", "Count", "", ""])
        row = style.write_data_rows(ws, row, [
            ["GEB list entries", len(scope.membership)],
            ["GEB list entries never matched", scope.unmatched_members],
        ])
        row += 1

    style.finalize_sheet(ws, freeze="A2", widths={"A": 40, "B": 14, "C": 14, "D": 14})


def _write_unmapped_regions(ws, row, frame):
    """The source region values no group could be resolved for.

    Named with their counts rather than swept into an "Unmapped" row, because
    this listing is how `regions.py`'s tables are meant to grow: a value that
    only ever appears inside a bucket total never gets fixed. Printed exactly as
    the source wrote it -- that is what somebody has to search for.
    """
    unmapped = regions.unmapped_values(frame.get("region", []))
    row = style.write_section_header(ws, row, "REGION VALUES NOT YET MAPPED", 4)
    row = style.write_header_row(ws, row, ["Value", "Activities", "", ""])
    if not unmapped:
        return style.write_data_rows(
            ws, row, [["Every region value resolves to a group", 0]])
    for value, count in unmapped:
        style.write_data_rows(ws, row, [[value, count]])
        row += 1
    return row


def _quarter_label(quarter):
    return f"Q{quarter[1]} {quarter[0]}"


def build_audience(wb, scope, config):
    """Audience size and GEB/GEB-1 involvement -- two of the three criteria the
    whole report is built around.

    The most interesting figure here is the share per division: involvement as
    a share of *that division's own volume*, not of the portfolio. A large
    division with many such activities may still be using that access less than
    a small one -- the portfolio-wide share would hide that.

    The last block names the people. It is the one place the report answers
    "whose activity is this", which the yes/no columns elsewhere cannot.
    """
    # Not "Audience & GEB/GEB-1": Excel forbids "/" in a sheet name. The tab
    # covers both leadership source fields anyway, so it is named for the pair.
    ws = wb.create_sheet("Audience & leadership")
    frame = scope.frame
    if frame.empty or "audience_band" not in frame.columns:
        style.note_missing(ws, "No audience data available (audience column missing)")
        style.finalize_sheet(ws, freeze="A2")
        return

    quarters = sorted({q for q in frame["_quarter"] if q is not None})
    headers = ["Audience band"] + [_quarter_label(q) for q in quarters] + ["Total", "% of total"]
    total_col = len(headers) - 1
    share_col = len(headers)

    row = style.write_section_header(ws, 1, "AUDIENCE BAND BY QUARTER", len(headers))
    row = style.write_header_row(ws, row, headers)
    first_row = row
    for band in list(AUDIENCE_BANDS) + [BAND_UNKNOWN]:
        counts = [int(((frame["audience_band"] == band) & (frame["_quarter"] == q)).sum())
                  for q in quarters]
        style.write_data_rows(ws, row, [[band] + counts])
        span = f"{get_column_letter(2)}{row}:{get_column_letter(total_col - 1)}{row}"
        style.write_formula(ws, row, total_col, f"=SUM({span})", fmt=style.NUM_FMT_INT)
        row += 1
    total_row = row
    for col in range(2, total_col + 1):
        letter = get_column_letter(col)
        style.write_formula(ws, total_row, col, f"=SUM({letter}{first_row}:{letter}{row - 1})",
                            fmt=style.NUM_FMT_INT, fill=style.TOTAL_FILL, bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = style.TOTAL_FONT
    total_letter = get_column_letter(total_col)
    for value_row in range(first_row, total_row):
        style.write_formula(
            ws, value_row, share_col,
            f"=IF({total_letter}${total_row}=0,0,"
            f"{total_letter}{value_row}/{total_letter}${total_row})",
            fmt=style.NUM_FMT_PCT)
    style.write_formula(ws, total_row, share_col, "=1", fmt=style.NUM_FMT_PCT,
                        fill=style.TOTAL_FILL, bold=True)
    row = total_row + 2

    row = style.write_section_header(ws, row, "LARGE AUDIENCE BY MONTH", 4)
    row = style.write_header_row(ws, row, ["Month", "Large audience", "All activities",
                                           "Share of the month"])

    def _week_month(index):
        # week_index is None for rows without a placeable week; on the
        # in-scope frame that never happens (every surviving row's start day
        # falls inside the grid), but a stray NaN from a mixed-dtype column
        # is guarded against rather than assumed away.
        if index is None or index != index:
            return None
        return scope.grid.month_of(scope.grid.weeks[int(index)])

    # Computed once per row up front, not once per row per month: the
    # reference version recomputed `month_of` inside an `.apply()` lambda for
    # every month, which is O(rows x months) for no benefit.
    week_months = frame["week_index"].map(_week_month)
    months = sorted({m for m in week_months if m is not None})
    for month in months:
        in_month = week_months == month
        large = int((in_month & frame["audience_band"].isin(LARGE_AUDIENCE_BANDS)).sum())
        style.write_data_rows(ws, row, [
            [f"{month[0]}-{month[1]:02d}", large, int(in_month.sum())]])
        style.write_formula(ws, row, 4, f"=IF(C{row}=0,0,B{row}/C{row})",
                            fmt=style.NUM_FMT_PCT)
        row += 1
    row += 1

    row = style.write_section_header(ws, row, "GEB/GEB-1 INVOLVEMENT BY QUARTER", 4)
    row = style.write_header_row(ws, row, ["Quarter", "With GEB/GEB-1", "All activities",
                                           "Share of the quarter"])
    for quarter in quarters:
        in_quarter = frame["_quarter"] == quarter
        style.write_data_rows(ws, row, [
            [_quarter_label(quarter),
             int((in_quarter & frame["has_executives"]).sum()),
             int(in_quarter.sum())]])
        style.write_formula(ws, row, 4, f"=IF(C{row}=0,0,B{row}/C{row})",
                            fmt=style.NUM_FMT_PCT)
        row += 1
    row += 1

    row = style.write_section_header(ws, row, "GEB/GEB-1 INVOLVEMENT BY DIVISION", 4)
    row = style.write_header_row(ws, row, ["Division", "With GEB/GEB-1",
                                           "All activities", "Share of the division"])
    divisions = {}
    for index, activity in frame.iterrows():
        for name in (split_multi(activity.get("business_division")) or ["Not specified"]):
            divisions.setdefault(name, []).append(index)
    for name in sorted(divisions):
        subset = frame.loc[divisions[name]]
        style.write_data_rows(ws, row, [
            [name, int(subset["has_executives"].sum()), len(subset)]])
        style.write_formula(ws, row, 4, f"=IF(C{row}=0,0,B{row}/C{row})",
                            fmt=style.NUM_FMT_PCT)
        row += 1
    # No TOTAL row here: an activity naming two divisions is counted in both
    # rows (see GLOSSARY_SECTIONS' "Overlap" entry), so a vertical SUM would
    # print a number larger than the portfolio, as if it were a true total.
    row += 1

    if scope.membership is None:
        row = _write_people_block(
            ws, row, frame, "executives",
            "ACTIVITIES BY GEB/GEB-1 MEMBER", "GEB/GEB-1 member",
            "All activities with GEB/GEB-1",
            "No GEB/GEB-1 member named on any in-scope activity")
        row += 1
    else:
        # One block per level, never a level block beside the combined one:
        # the same person would appear twice with the same count.
        row = _write_people_block(
            ws, row, frame, "executives_geb",
            "ACTIVITIES BY GEB MEMBER", "GEB member",
            "All activities with GEB",
            "No GEB member named on any in-scope activity")
        row += 1
        row = _write_people_block(
            ws, row, frame, "executives_geb1",
            "ACTIVITIES BY GEB-1 MEMBER", "GEB-1 member",
            "All activities with GEB-1",
            "No GEB-1 member named on any in-scope activity")
        row += 1

    row = _write_people_block(
        ws, row, frame, "senior_executives",
        "ACTIVITIES BY OTHER SENIOR EXECUTIVE", "Senior executive",
        "All activities with other senior executives",
        "No other senior executive named on any in-scope activity")

    style.finalize_sheet(ws, freeze="B3", widths={"A": 26})


def _write_people_block(ws, row, frame, field, title, noun, total_label, empty_note):
    """One block of person -> activity count -> share, for one people field.

    Both leadership fields render through here so they cannot drift apart:
    same ordering (volume, then name), same denominator rule, same handling of
    an empty field. Only the wording differs.

    No TOTAL row. An activity naming two people counts under both, so the
    shares can legitimately add up to past 100% -- the honest reading of shared
    ownership. A forced 100% would have to pick a winner.
    """
    row = style.write_section_header(ws, row, title, 3)
    row = style.write_header_row(ws, row, [noun, "Activities", "Share of these"])

    people = {}
    for index, activity in frame.iterrows():
        for name in split_people(activity.get(field)):
            people.setdefault(name, []).append(index)

    # Activities with at least one name, counted once each however many names
    # they carry -- derived from the very rows listed below rather than from a
    # separate column, so the denominator and its members cannot disagree.
    involved = len({index for indices in people.values() for index in indices})

    # A real cell, not a literal baked into each formula, so a reader can click
    # any share and follow it to the number it divides by.
    style.write_data_rows(ws, row, [[total_label, involved]])
    ws.cell(row=row, column=1).font = style.TOTAL_FONT
    for col in (1, 2):
        ws.cell(row=row, column=col).fill = style.TOTAL_FILL
    involved_row = row
    row += 1

    if not people:
        style.write_data_rows(ws, row, [[empty_note]])
        return row + 1

    for name in sorted(people, key=lambda n: (-len(people[n]), n)):
        style.write_data_rows(ws, row, [[name, len(people[name])]])
        style.write_formula(
            ws, row, 3,
            f"=IF($B${involved_row}=0,0,B{row}/$B${involved_row})",
            fmt=style.NUM_FMT_PCT)
        row += 1
    return row


GLOSSARY_SECTIONS = (
    ("SCOPE", (
        ("In scope", "Activities that match the criteria listed at the top of the "
                     "Executive Summary."),
    )),
    ("DIMENSIONS", (
        ("Region", "Americas, APAC, EMEA, Switzerland or Global, resolved from the "
                   "source value. Switzerland is its own group."),
        ("Country", "The country behind the source value. Cities roll up into "
                    "theirs; a region-only value has none."),
        ("Overlap", "An activity naming two divisions counts in both, so those blocks "
                    "add up to more than the total."),
        # The calendar and the people lists answer different questions about the
        # same field, and their totals therefore differ: the calendar groups the
        # whole portfolio, so an activity naming nobody needs a bucket, while a
        # roster of people has no row for "nobody". Both are right; a reader
        # comparing the two headings without this line sees a contradiction.
        ("Not specified", "Activities with no value in that field. The calendar gives "
                          "them a row; the people lists leave them out."),
    )),
    ("MEASURES", (
        ("Audience band", "The size band of the target audience."),
        # The caveat, not the pointer, earns the scarce characters here: a reader
        # who reads this as "GEB" draws a wrong conclusion, whereas one who does
        # not know where the names are listed merely has to look.
        ("GEB/GEB-1", "At least one person at GEB or GEB-1 level. The source mixes "
                      "the two, so not everyone named is on the GEB."),
        ("Other senior executives", "A senior executive from the source's separate "
                                    "other-executives field. Counted separately."),
        ("Lead time", "Days from creating the record to the activity's start."),
        ("Planning completeness", "Share of the required fields that are filled in."),
        ("Weekly counts", "Each activity counts once, in the week it starts."),
        # Says what a pack is. The old wording ("Not used as a grouping
        # dimension") described a limitation that the pack file removed, and
        # a definition that describes the tool rather than the thing goes
        # stale the moment the tool changes.
        ("Packs", "A communications pack: activities grouped around one "
                  "communication objective, with its own lead and period."),
    )),
)


# Added to MEASURES only when a membership list is in play -- defining terms
# the workbook never prints would be its own small lie.
#
# GEB alone, on request 2026-08-06. The GEB-1 entry and GEB's "everyone else in
# the field is GEB-1" clause both went; the sheets still label a GEB-1 block, so
# that heading now stands without a definition. Deliberate, and not a gap to
# quietly fill: restoring either is a product decision.
GEB_SPLIT_TERMS = (
    ("GEB", "A person named on the GEB list."),
)


def _glossary_sections(scope):
    """The Glossary's sections, with the GEB term present only when a
    membership list is in play.

    Assembled rather than branched at the write site so the section order,
    the widths and the wrapping stay in one place.
    """
    if scope.membership is None:
        return GLOSSARY_SECTIONS
    sections = []
    for title, terms in GLOSSARY_SECTIONS:
        if title == "MEASURES":
            expanded = []
            for entry in terms:
                expanded.append(entry)
                if entry[0] == "GEB/GEB-1":
                    expanded.extend(GEB_SPLIT_TERMS)
            terms = tuple(expanded)
        sections.append((title, terms))
    return tuple(sections)


def _crosstab_block(ws, row, quarters, frame, title, field, sort_key=None):
    """One label x quarter table with a Total column.

    `sort_key` orders the label rows -- plain alphabetical by default, which
    is wrong for priority and right for everything else.

    Rows overlap for multi-valued fields (channel, division): an activity
    naming two channels counts in both of that field's rows. No TOTAL row is
    written across the label rows for the same reason the Audience sheet's
    division block omits one -- a vertical SUM would print a number larger
    than the portfolio, as if it were a true total.

    There is deliberately no quarter-to-quarter difference column, and adding
    one back needs a product decision rather than a fix. The one that shipped
    here compared the first quarter in scope against the last full one, which
    sets a quarter that is finished against a quarter still being filled in:
    every row read strongly negative, and what it measured was how far ahead
    the planning had got, not how the mix had moved. Removed 2026-08-06.
    """
    if not quarters:
        style.write_section_header(ws, row, f"{title} — no data", 3)
        return row + 2

    headers = ["Value"] + [_quarter_label(q) for q in quarters] + ["Total"]
    total_col = len(headers)

    row = style.write_section_header(ws, row, title, len(headers))
    row = style.write_header_row(ws, row, headers)

    values = {}
    for index, activity in frame.iterrows():
        for name in (split_multi(activity.get(field)) or ["Not specified"]):
            values.setdefault(name, []).append(index)

    first_letter = get_column_letter(2)
    last_letter = get_column_letter(total_col - 1)
    for name in sorted(values, key=sort_key):
        subset = frame.loc[values[name]]
        counts = [int((subset["_quarter"] == q).sum()) for q in quarters]
        style.write_data_rows(ws, row, [[name] + counts])
        span = f"{first_letter}{row}:{last_letter}{row}"
        style.write_formula(ws, row, total_col, f"=SUM({span})", fmt=style.NUM_FMT_INT)
        row += 1
    return row + 1


def _priority_sort_key(name):
    """Most urgent first. Alphabetical order is actively misleading here.

    Two priority vocabularies are live at once -- the source system's numbered
    labels ("1 - ...", where 1 is most urgent) and the studio's words -- so an
    alphabetical sort interleaves them and puts Low above Medium. Matching only
    the words has already produced a metric reading zero against thousands of
    records (see the Mix sheet's note in the design document), which is why
    this goes through the same rank function the studio uses. The name is the
    tiebreaker so the order is stable within a rank.
    """
    return (-priority_rank(name), name)


def build_mix(wb, scope, config):
    """Channel, priority and internal/external mix over the quarters in scope,
    plus lead time by division -- how far ahead planning actually happens.
    """
    ws = wb.create_sheet("Mix & Lead Time")
    frame = scope.frame
    if frame.empty:
        style.note_missing(ws, "No activities in scope for the configured criteria")
        style.finalize_sheet(ws, freeze="A2")
        return

    quarters = sorted({q for q in frame["_quarter"] if q is not None})

    row = _crosstab_block(ws, 1, quarters, frame, "CHANNEL BY QUARTER", "channel")
    row = _crosstab_block(ws, row, quarters, frame, "PRIORITY BY QUARTER", "priority",
                          sort_key=_priority_sort_key)
    row = _crosstab_block(ws, row, quarters, frame,
                          "INTERNAL VS EXTERNAL BY QUARTER", "source_type")

    row = style.write_section_header(ws, row, "LEAD TIME BY DIVISION", 6)
    row = style.write_header_row(ws, row, [
        "Division", "Measurable", "Median days",
        f"Under {SHORT_NOTICE_DAYS} days", "Share short notice", "Min / max days"])
    divisions = {}
    for index, activity in frame.iterrows():
        for name in (split_multi(activity.get("business_division")) or ["Not specified"]):
            divisions.setdefault(name, []).append(index)
    for name in sorted(divisions):
        stats = metrics.lead_time_stats(frame.loc[divisions[name]])
        span = "—" if stats["min_days"] is None else f"{stats['min_days']} / {stats['max_days']}"
        style.write_data_rows(ws, row, [[
            name, stats["counted"],
            stats["median_days"] if stats["median_days"] is not None else "—",
            stats["short_notice"], None, span]])
        style.write_formula(ws, row, 5, f"=IF(B{row}=0,0,D{row}/B{row})",
                            fmt=style.NUM_FMT_PCT)
        row += 1

    style.finalize_sheet(ws, freeze="B3", widths={"A": 30})


ACTIVITY_COLUMNS = (
    ("tracking_id", "Tracking ID"),
    ("activity_name", "Activity"),
    ("source_type", "Type"),
    ("channel", "Channel"),
    ("start_date", "Start"),
    ("end_date", "End"),
    ("_iso_week", "ISO week"),
    ("_quarter_label", "Quarter"),
    ("priority", "Priority"),
    ("lead", "Lead"),
    ("lead_team", "Lead team"),
    ("target_audience", "Target audience"),
    ("audience_band", "Audience band"),
    ("business_division", "Divisions"),
    ("region", "Regions (as recorded)"),
    ("region_group", "Region"),
    ("country", "Country"),
    ("_executives", "GEB/GEB-1 involved"),
    ("executives", "GEB/GEB-1 members"),
    ("senior_executives", "Other senior executives"),
    # The name beside the number. `communication_pack` is mapped and
    # lookup-parsed exactly like every other reference field; it was simply
    # never exported, so the pack has always shown the identifier alone. A
    # reader asks about "Pack one", never about "CP-100". The identifier
    # stays because it is what `07-packs.csv` joins on, and because a pack
    # name is not unique the way a key is.
    ("communication_pack", "Pack"),
    ("communication_pack_cpid", "Pack ID"),
    ("campaign", "Campaign"),
    ("strategic_objectives", "Communications pillars"),
    ("completeness", "Completeness %"),
    ("lead_time_days", "Lead time (days)"),
    ("is_archived", "Archived"),
)


def build_activities(wb, scope, config):
    """Every in-scope activity, one row each, raw fields and derived ones
    side by side -- the audit trail every other sheet's figures trace back to.
    """
    ws = wb.create_sheet("Activities")
    frame = scope.frame
    style.write_header_row(ws, 1, [header for _, header in ACTIVITY_COLUMNS])
    if frame.empty:
        style.finalize_sheet(ws, freeze="A2")
        return

    rows = []
    for _, activity in frame.iterrows():
        index = activity["week_index"]
        week = scope.grid.weeks[int(index)] if index == index and index is not None else None
        quarter = activity["_quarter"]
        values = []
        for field, _ in ACTIVITY_COLUMNS:
            if field == "_iso_week":
                values.append(f"{week.iso_year}-{week.label}" if week else "")
            elif field == "_quarter_label":
                values.append(_quarter_label(quarter) if quarter else "")
            elif field == "_executives":
                values.append("Yes" if activity["has_executives"] else "No")
            elif field in ("start_date", "end_date"):
                value = activity.get(field)
                values.append(value.date() if hasattr(value, "date") and value == value
                              else None)
            elif field == "lead_time_days":
                value = activity.get(field)
                values.append(int(value) if value == value and value is not None else None)
            else:
                value = activity.get(field)
                values.append("" if value is None or value != value else value)
        rows.append(values)

    date_columns = {
        i + 1: style.NUM_FMT_DATE
        for i, (field, _) in enumerate(ACTIVITY_COLUMNS)
        if field in ("start_date", "end_date")
    }
    style.write_data_rows(ws, 2, rows, fmt_map=date_columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ACTIVITY_COLUMNS))}{len(rows) + 1}"
    style.finalize_sheet(ws, freeze="A2")


def build_glossary(wb, scope, config):
    ws = wb.create_sheet("Glossary")
    ws.sheet_view.showGridLines = False
    row = 1
    for title, terms in _glossary_sections(scope):
        row = style.write_section_header(ws, row, title, 2)
        for term, definition in terms:
            style.write_kpi_row(ws, row, term, None)
            cell = ws.cell(row=row, column=2, value=definition)
            cell.font = style.BODY_FONT
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal, wrap_text=True, vertical="top")
            row += 1
        row += 1

    if scope.skipped_completeness_fields:
        row = style.write_section_header(ws, row, "FIELDS NOT IN THIS EXPORT", 2)
        for name in scope.skipped_completeness_fields:
            row = style.write_kpi_row(
                ws, row, name,
                "Not in the export, so not counted.")

    style.finalize_sheet(ws, freeze=None, widths={"A": 28, "B": 90})
