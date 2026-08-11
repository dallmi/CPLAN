"""The calendar matrix.

Rows are planning dimensions, columns are quarters that expand into months that
expand into ISO weeks. Both axes are Excel outlines, so the sheet opens as a
handful of quarter columns against a handful of block rows and expands on click.

Two rules keep the arithmetic honest:

* Horizontal aggregation is ALWAYS a formula -- a month is the sum of its week
  cells, a quarter the sum of its months, the total the sum of its quarters. The
  sheet cannot contradict itself and the reader can click any figure. This holds
  for every row, including block headers: week cells are always literal counts,
  never a formula, in every row of the sheet.
* Vertical aggregation is a formula only where it is valid, and only in the
  Total (summary) column. The audience bands partition the portfolio, so that
  block header's Total cell carries a genuine SUM down the member rows -- a
  second, independently auditable route to the same number as the horizontal
  route. The division and region blocks overlap -- an activity naming two
  divisions appears twice -- so their headers' Total cell carries a distinct
  count computed here, and say so in the label. A SUM there would print a bold
  number larger than the portfolio.
"""

from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from pipeline.report import regions, style
from pipeline.report.config import AUDIENCE_BAND_ORDER, EXECUTIVES_SPLIT, FIELD_TITLES
from pipeline.report.derive import priority_rank, split_multi, split_semicolon

SHEET_NAME = "Calendar"
LABEL_COL = 1
TOTAL_COL = 2
FIRST_GRID_COL = 3
FIRST_DATA_ROW = 3

NOT_SPECIFIED = "Not specified"

# FIELD_TITLES itself now lives in config.py: `ReportConfig.describe()` needs
# the same reader-facing names for its "Breakdown dimensions" row, and two
# copies of this map would only be able to disagree, never help.

# Most blocks read best alphabetically. The region groups do not: they have a
# natural reading order, and Global -- the largest value in the source -- would
# otherwise sit between EMEA and Switzerland.
#
# Priority is not here. It used to be, listing the four bucket names the pack
# folded its values onto; the pack now carries the source's own labels, and
# there is no fixed list of those to enumerate. `_sort_key` ranks it instead.
# A map naming values nothing produces any more would order nothing and read
# as though it did.
FIELD_ORDER = {
    "region_group": {name: i for i, name in enumerate(regions.GROUP_ORDER)},
}


def _sort_key(field, name):
    # Priority has no fixed list to order by: two vocabularies are live at
    # once -- the source's numbered labels and the studio's words -- and both
    # appear verbatim rather than folded onto a shared four. So it orders by
    # the rank both of them yield, most urgent first, exactly as the
    # workbook's Mix sheet does. Name breaks ties so the order is stable.
    if field == "priority":
        return (name == NOT_SPECIFIED, -priority_rank(name), name)
    order = FIELD_ORDER.get(field)
    if order is None:
        return (name == NOT_SPECIFIED, name)
    return (name == NOT_SPECIFIED, order.get(name, len(order)), name)


# People fields split on the semicolon only: a display name contains a comma
# ("Last, First"), so the generic splitter would turn one person into two rows.
PEOPLE_FIELDS = frozenset({
    "executives", "senior_executives", "executives_geb", "executives_geb1"})

# The two fields a membership list derives from `executives`, and only those:
# see the breakdown-fields loop in `build_calendar` for why they skip the
# generic "Not specified" catch-all that every other breakdown field gets.
# Sourced from `config.EXECUTIVES_SPLIT` rather than named again here, so this
# set and `report_calendar.py`'s swap cannot silently drift apart.
SPLIT_FIELDS = frozenset(EXECUTIVES_SPLIT)

# Breakdown fields that name exactly one value per activity, and so PARTITION
# the portfolio -- their rows genuinely sum to the total, the way the audience
# bands do. All three are fields `agent_pack.pack_config` adds; every field the
# workbook itself offers (division, region, country, executives) is multi-select
# at the source and stays overlapping, which is why the Calendar sheet titles
# each of THEIR blocks "multiple values possible" unconditionally rather than
# checking whether any particular run's data happens to use it. Declared here
# rather than inferred from one run's rows for the same reason: an activity
# that merely doesn't name a second division this year does not make the field
# a partition, and a block's overlap sentence has to hold regardless of what
# one dataset does.
#
# Declared HERE, next to the splitter, rather than in `agent_pack` where the
# `overlaps` column is written: the partition claim and the separator that
# makes it true are one decision. Two sets, one per module, could only ever
# disagree -- the same reason `SPLIT_FIELDS` reads `EXECUTIVES_SPLIT` instead
# of naming those columns a second time.
PARTITION_BREAKDOWN_FIELDS = frozenset({"priority", "lead_team", "source_type"})


def _split_for(field, value):
    # A partitioning field splits on the semicolon for the same reason a people
    # field does, and the stakes are higher. Its one value is a label somebody
    # typed, and a team named after two disciplines carries a comma inside its
    # own name; the generic splitter reads that as two teams, files the
    # activity under both, and the block that tells every reader `overlaps=no`
    # then sums to more than the portfolio. The semicolon stays a separator, so
    # a value that really does name two teams still reaches
    # `agent_pack.iter_blocks`'s guard rather than quietly becoming one bucket
    # labelled "A; B".
    if field in PEOPLE_FIELDS or field in PARTITION_BREAKDOWN_FIELDS:
        return split_semicolon(value)
    return split_multi(value)


def _column_positions(columns):
    """Map each grid column to its sheet column index, and group the children."""
    positions = {}
    for offset, column in enumerate(columns):
        positions[(column.kind, column.key)] = FIRST_GRID_COL + offset
    return positions


def _children(columns, grid):
    """For each month column, its week columns; for each quarter, its months."""
    month_weeks = {}
    quarter_months = {}
    current_quarter = None
    current_month = None
    for column in columns:
        if column.kind == "quarter":
            current_quarter = column.key
            quarter_months.setdefault(current_quarter, [])
        elif column.kind == "month":
            current_month = column.key
            month_weeks.setdefault(current_month, [])
            quarter_months[current_quarter].append(current_month)
        else:
            month_weeks[current_month].append(column.key)
    return month_weeks, quarter_months


def build_row_plan(columns, positions, month_weeks, quarter_months):
    """Everything about the grid's columns that does not depend on the row.

    The per-row writer used to rebuild each month's and quarter's child-column
    letters from scratch, for every row in the sheet. Identical work, thousands
    of times: on a wide grid that is the difference between a report and a
    hang. Computed once here, and the row number is the only thing left to
    substitute.
    """
    weeks, sums = [], []
    for column in columns:
        col = positions[(column.kind, column.key)]
        if column.kind == "week":
            weeks.append((column.key, col))
            continue
        if column.kind == "month":
            child_keys = [("week", key) for key in month_weeks[column.key]]
        else:
            child_keys = [("month", key) for key in quarter_months[column.key]]
        letters = [get_column_letter(positions[key]) for key in child_keys]
        sums.append((col, "=SUM(" + ",".join(f"{letter}{{row}}" for letter in letters) + ")"))

    quarter_letters = [
        get_column_letter(positions[("quarter", column.key)])
        for column in columns if column.kind == "quarter"
    ]
    total = "=SUM(" + ",".join(f"{letter}{{row}}" for letter in quarter_letters) + ")"
    return weeks, sums, total


def _write_grid_row(ws, row, counts, plan, bold=False):
    """Literal week counts, SUM formulas everywhere else (month, quarter, Total).

    This is the ONLY way grid cells are populated -- every row in the sheet,
    including block header rows, gets its week/month/quarter/Total cells this
    way. A block header's Total cell may be overwritten afterwards (see
    `_finish_partition_header` and `_finish_distinct_count_header`), but its
    week/month/quarter cells always come from here, so the horizontal identity
    (month = SUM of its weeks, quarter = SUM of its months) never breaks.

    An empty week in an ordinary row is left untouched -- no cell, no style.
    Nearly every cell in the grid is empty (one activity occupies one week out
    of hundreds), and styling them was 87% of the sheet's build time: openpyxl
    re-hashes the whole border object on every assignment. Excel draws its own
    gridlines there, and SUM reads an absent cell as zero, so nothing is lost.
    Header rows still fill theirs, because their shading is the block's outline.
    """
    weeks, sums, total = plan
    fill = style.TOTAL_FILL if bold else None
    for key, col in weeks:
        value = counts.get(key, 0)
        if not value and not bold:
            continue
        cell = ws.cell(row=row, column=col, value=value or None)
        cell.border = style.THIN_BORDER
        if value:
            cell.number_format = style.NUM_FMT_INT
        if bold:
            cell.font = style.TOTAL_FONT
            cell.fill = style.TOTAL_FILL

    for col, template in sums:
        style.write_formula(ws, row, col, template.format(row=row),
                            fmt=style.NUM_FMT_INT, fill=fill, bold=bold)

    total_formula = total.format(row=row)
    style.write_formula(ws, row, TOTAL_COL, total_formula, fmt=style.NUM_FMT_INT,
                        fill=fill, bold=bold)


def _counts(frame, grid):
    """Week key -> number of activities starting in that week.

    Every row in `frame` is assumed to carry a resolved (non-NaN) `week_index`.
    That precondition is real, not defensive fiction: `build_scope` filters a
    row into the grid's own date window (dropping "no start date" and "date
    window" rows) *before* it computes `week_index`, so every surviving row's
    start day falls inside the grid and resolves to a real week. It is
    asserted once, on entry, in `build_calendar` -- so it is not re-checked
    here. A NaN reaching this loop would mean that upstream invariant broke,
    and a loud crash close to the cause beats a silent undercount far from
    it (see the sibling detail-row loop in `build_calendar`, which also
    relies on the same guarantee without a local guard).
    """
    counts = {}
    for index in frame["week_index"]:
        key = grid.weeks[int(index)].key
        counts[key] = counts.get(key, 0) + 1
    return counts


def _mark_collapsed(dimension):
    """Tell the reader's spreadsheet that the group under this summary is shut.

    `hidden` alone takes the rows and columns out of view but says nothing
    about the outline. Excel draws the +/- control and its state from
    `collapsed` on the SUMMARY row/column -- which, with summaryBelow and
    summaryRight both off, is the one immediately above or to the left of the
    group. Without it the detail is invisible and there is no control to click:
    the sheet reads as almost empty, which is exactly how it was first reported
    from a real Excel.
    """
    dimension.collapsed = True


def _detail_label(activity):
    """An activity's row label, naming the GEB/GEB-1 people where there are any.

    The BY GEB/GEB-1 block already answers "which activities are this person's".
    This answers the reverse from inside every other block: expanding a division
    shows at a glance which of its activities carry that involvement, and whose.
    Text rather than colour -- a colour can say "somebody", never "who", and the
    Total column's data bars already own the sheet's one visual channel.
    """
    name = activity.get("activity_name") or "Untitled"
    executives = activity.get("executives") or ""
    return f"  {name} — {executives}" if executives else f"  {name}"


def _label_cell(ws, row, text, level, bold=False, hidden=False):
    cell = ws.cell(row=row, column=LABEL_COL, value=style.safe(text))
    cell.border = style.THIN_BORDER
    if bold:
        cell.font = style.TOTAL_FONT
        cell.fill = style.TOTAL_FILL
    ws.row_dimensions[row].outline_level = level
    if hidden:
        ws.row_dimensions[row].hidden = True


def build_calendar(wb, scope, config):
    ws = wb.create_sheet(SHEET_NAME)
    grid = scope.grid
    columns = grid.columns()
    positions = _column_positions(columns)
    month_weeks, quarter_months = _children(columns, grid)
    plan = build_row_plan(columns, positions, month_weeks, quarter_months)

    ws.sheet_properties.outlinePr.summaryRight = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    # How deep the outline goes, on each axis. Excel sizes and draws the
    # outline gutter from these; at 0 there is no gutter, so a sheet that opens
    # collapsed offers nothing to expand it with.
    #
    # The two axes need different handling, which is not obvious and cost a
    # debugging session. The row value on `sheet_format` is written through
    # untouched. The column value is NOT: openpyxl's worksheet writer does
    # `sheet_format.outlineLevelCol = column_dimensions.max_outline`
    # unconditionally on save, and `max_outline` is a plain attribute that only
    # `DimensionHolder.group()` ever sets -- assigning `outline_level` on the
    # individual columns, as this builder does, leaves it None. So setting
    # `sheet_format.outlineLevelCol` here would be silently discarded, and the
    # column outline has to be declared on the holder instead.
    depth = max((column.level for column in columns), default=0)
    ws.sheet_format.outlineLevelRow = 2
    ws.column_dimensions.max_outline = depth

    # --- header -------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=LABEL_COL, end_row=2, end_column=LABEL_COL)
    ws.merge_cells(start_row=1, start_column=TOTAL_COL, end_row=2, end_column=TOTAL_COL)
    style.write_header_row(ws, 1, ["Scope / activity", "Total"])
    # `write_header_row` only styles row 1; the merge means A2/B2 are blank
    # and unstyled underneath it unless we style them too, which would show
    # as a gap in the header band (no fill, no border) where every grid
    # column has both rows filled and bordered.
    for col in (LABEL_COL, TOTAL_COL):
        under_cell = ws.cell(row=2, column=col)
        under_cell.fill = style.HEADER_FILL
        under_cell.border = style.THIN_BORDER
    grid_widths = {}
    for column in columns:
        col = positions[(column.kind, column.key)]
        style.write_header_row(ws, 1, [column.label], col_start=col)
        style.write_header_row(ws, 2, [column.sublabel], col_start=col)
        letter = get_column_letter(col)
        ws.column_dimensions[letter].outline_level = column.level
        ws.column_dimensions[letter].hidden = column.level > 0
        # A quarter column summarises the month group to its right, a month
        # column its week group. Both open shut, so both carry the flag.
        if column.kind in ("quarter", "month"):
            _mark_collapsed(ws.column_dimensions[letter])
        # NOT set directly on column_dimensions here: `finalize_sheet` calls
        # `auto_fit_columns`, which unconditionally overwrites every column's
        # width from its longest cell content (including raw formula text,
        # which is never what's displayed). Only the `widths=` dict passed to
        # `finalize_sheet` survives that pass, so the intended widths are
        # collected here and applied there instead.
        grid_widths[letter] = 11 if column.kind == "week" else 13

    if scope.frame.empty:
        cell = ws.cell(row=FIRST_DATA_ROW, column=LABEL_COL,
                        value="No activities in scope for the configured criteria")
        cell.font = style.SUB_FONT
        style.finalize_sheet(ws, freeze="C3", widths={**grid_widths, "A": 52, "B": 12})
        return

    # `_counts` (and the detail-row loop below) assume every row's week_index
    # already resolved -- true by construction (see `_counts`'s docstring),
    # asserted here once so a broken upstream invariant fails loudly and close
    # to its cause instead of silently corrupting a header's arithmetic.
    assert scope.frame["week_index"].notna().all(), (
        "every row in scope.frame must carry a resolved week_index; "
        "build_scope is expected to guarantee this by construction"
    )

    row = FIRST_DATA_ROW
    bar_ranges = []

    # --- all activities -----------------------------------------------------
    _label_cell(ws, row, "ALL ACTIVITIES", level=0, bold=True)
    _write_grid_row(ws, row, _counts(scope.frame, grid), plan, bold=True)
    row += 1

    def write_value_row(label, subset, level, hidden):
        nonlocal row
        _label_cell(ws, row, label, level=level, hidden=hidden)
        _write_grid_row(ws, row, _counts(subset, grid), plan)
        value_row = row
        row += 1
        if config.detail_rows and not subset.empty:
            _mark_collapsed(ws.row_dimensions[value_row])
            ordered = subset.sort_values("start_day", kind="stable")
            for _, activity in ordered.iterrows():
                _label_cell(ws, row, _detail_label(activity),
                            level=level + 1, hidden=True)
                week_key = grid.weeks[int(activity["week_index"])].key
                _write_grid_row(ws, row, {week_key: 1}, plan)
                row += 1
        return value_row

    # --- audience: a partition, so its Total is a genuine SUM down the column
    # Every activity carries exactly one band, Unknown included, which is what
    # makes the vertical SUM below valid. The overlapping blocks that follow
    # cannot do this and say so in their own headers.
    _label_cell(ws, row, "BY AUDIENCE", level=0, bold=True)
    header_row = row
    row += 1
    member_rows = []
    for band in AUDIENCE_BAND_ORDER:
        subset = scope.frame[scope.frame["audience_band"] == band]
        if subset.empty:
            continue
        member_rows.append(write_value_row(band, subset, level=1, hidden=True))
    # Week/month/quarter cells stay the normal, honest way: literal weekly
    # counts over the whole (partitioned) scope, horizontal SUMs above them --
    # identical in shape to the ALL ACTIVITIES row, since a true partition's
    # per-week total is the same number either way. Only the Total (B) column
    # is deliberately written as a second, independent formula that sums the
    # member rows vertically, so a reader can audit that the audience bands
    # really do add back up to the portfolio.
    _write_grid_row(ws, header_row, _counts(scope.frame, grid), plan, bold=True)
    _finish_partition_header(ws, header_row, member_rows)
    if member_rows:
        _mark_collapsed(ws.row_dimensions[header_row])
    bar_ranges.append(member_rows)

    # --- breakdown fields: overlapping, so a distinct count -----------------
    for field in config.breakdown_fields:
        if field not in scope.frame.columns:
            continue
        title = f"BY {FIELD_TITLES.get(field, field.upper())} — multiple values possible"
        _label_cell(ws, row, title, level=0, bold=True)
        header_row = row
        row += 1
        values = {}
        for _, activity in scope.frame.iterrows():
            names = _split_for(field, activity.get(field))
            if not names:
                # A split field (see `SPLIT_FIELDS`) has no catch-all bucket of
                # its own: an activity with nothing on this side of the split
                # either has something on the other side or has no GEB/GEB-1
                # people at all, and either way the un-split combined field
                # already accounts for it as `Not specified` when no
                # membership is loaded. Giving GEB and GEB-1 each their own
                # catch-all row would count that activity in both blocks.
                if field in SPLIT_FIELDS:
                    continue
                names = [NOT_SPECIFIED]
            for name in names:
                values.setdefault(name, []).append(activity.name)
        member_rows = []
        for name in sorted(values, key=lambda n: _sort_key(field, n)):
            subset = scope.frame.loc[values[name]]
            member_rows.append(write_value_row(name, subset, level=1, hidden=True))
        # Week/month/quarter cells: the same true, non-overlapping distinct
        # count as ALL ACTIVITIES (an activity tagged with two divisions is
        # still counted once here) -- true for any field with a catch-all
        # bucket, since every activity then lands in some member row and the
        # union is the whole scope. A split field carries no catch-all, so its
        # header instead counts only the activities that actually landed in
        # one of ITS OWN member rows; otherwise the GEB and GEB-1 headers
        # would each separately claim the whole scope and no longer sum back
        # to the combined field's own figure. Only the Total column is
        # overwritten with a literal below -- never a SUM -- because summing
        # the member rows vertically would double-count activities that
        # appear in more than one division/region (or, for a split field, in
        # both GEB and GEB-1). That literal is derived from the SAME `counts`
        # dict as the row's own week cells (`sum(counts.values())`), keeping
        # the header structurally in agreement with the cells printed beside
        # it rather than incidentally so.
        covered = sorted({idx for indices in values.values() for idx in indices})
        counts = _counts(scope.frame.loc[covered], grid)
        _write_grid_row(ws, header_row, counts, plan, bold=True)
        _finish_distinct_count_header(ws, header_row, sum(counts.values()))
        if member_rows:
            _mark_collapsed(ws.row_dimensions[header_row])
        bar_ranges.append(member_rows)

    for member_rows in bar_ranges:
        if not member_rows:
            continue
        sqref = " ".join(f"B{r}" for r in member_rows)
        ws.conditional_formatting.add(sqref, DataBarRule(
            start_type="num", start_value=0, end_type="max",
            color=style.GRAY_IV, showValue=True))

    style.finalize_sheet(ws, freeze="C3", widths={**grid_widths, "A": 52, "B": 12})


def _finish_partition_header(ws, header_row, member_rows):
    """Overwrite the Total cell with a genuine SUM down the member rows.

    Valid only for a partition: the audience block. Every activity lands in
    exactly one audience band, so summing the member rows' Total column gives
    back the same number as the horizontal route (summing this row's own
    quarters) -- the reader can click either path.
    """
    formula = "=SUM(" + ",".join(f"B{r}" for r in member_rows) + ")" \
        if member_rows else "=0"
    style.write_formula(ws, header_row, TOTAL_COL, formula, fmt=style.NUM_FMT_INT,
                        fill=style.TOTAL_FILL, bold=True)


def _finish_distinct_count_header(ws, header_row, total):
    """Overwrite the Total cell with a literal, never a SUM.

    Valid only where members overlap: the division and region blocks. An
    activity naming two divisions appears in two member rows, so a vertical
    SUM here would print a bold number larger than the portfolio. The true
    count is computed in Python instead and written as a plain literal.
    """
    cell = ws.cell(row=header_row, column=TOTAL_COL, value=total)
    cell.font = style.TOTAL_FONT
    cell.fill = style.TOTAL_FILL
    cell.number_format = style.NUM_FMT_INT
    cell.border = style.THIN_BORDER
