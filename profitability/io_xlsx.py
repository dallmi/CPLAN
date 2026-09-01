"""Excel (.xlsx) input/output for drivers, cost lines and allocation results.

Requires the optional dependency ``openpyxl`` (``pip install profitability[xlsx]``).
The expected column layout matches the CSV formats: a header row with
``driver,object,value`` for drivers and ``cost,amount,driver`` for costs.
Header matching is case-insensitive and ignores surrounding whitespace, extra
columns are ignored, and fully empty rows are skipped.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterator, Mapping, Optional, Union

from .allocation import AllocationRow, CostLine
from .distribution import Distribution
from .io_csv import COST_COLUMNS, DRIVER_COLUMNS

PathLike = Union[str, Path]


def _openpyxl():
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - exercised without the extra
        raise ImportError(
            "reading/writing .xlsx files needs the optional dependency 'openpyxl'; "
            "install it with: pip install openpyxl"
        ) from error
    return openpyxl


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_decimal(value: object, *, path: PathLike, row: int, column: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{path}: row {row}: invalid {column} {value!r}")
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # str() gives the shortest text that round-trips, so 0.1 stays 0.1.
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        raise ValueError(f"{path}: row {row}: invalid {column} {value!r}") from None


def _iter_rows(
    path: PathLike, required: tuple[str, ...], sheet: Optional[str]
) -> Iterator[tuple[int, dict[str, object]]]:
    """Yield ``(row_number, {column: cell_value})`` for each non-empty row."""
    openpyxl = _openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is not None:
            if sheet not in workbook.sheetnames:
                known = ", ".join(workbook.sheetnames)
                raise ValueError(f"{path}: no sheet {sheet!r}; sheets: {known}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError(f"{path}: workbook has no active sheet")

        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError(f"{path}: sheet {worksheet.title!r} is empty") from None
        positions: dict[str, int] = {}
        for index, cell in enumerate(header):
            name = _cell_text(cell).lower()
            if name and name not in positions:
                positions[name] = index
        missing = [column for column in required if column not in positions]
        if missing:
            raise ValueError(
                f"{path}: sheet {worksheet.title!r}: missing column(s) "
                f"{', '.join(missing)}; expected columns {', '.join(required)}"
            )

        for row_number, cells in enumerate(rows, start=2):
            record = {
                column: (cells[index] if index < len(cells) else None)
                for column, index in positions.items()
            }
            if all(_cell_text(value) == "" for value in record.values()):
                continue
            yield row_number, record
    finally:
        workbook.close()


def read_drivers_xlsx(
    path: PathLike, *, sheet: Optional[str] = None, negatives: str = "error"
) -> dict[str, Distribution]:
    """Read driver values from an Excel sheet with columns ``driver,object,value``
    and return one normalised :class:`Distribution` per driver.
    """
    raw: dict[str, dict[str, Decimal]] = {}
    for row_number, record in _iter_rows(path, DRIVER_COLUMNS, sheet):
        driver = _cell_text(record["driver"])
        obj = _cell_text(record["object"])
        if not driver or not obj:
            raise ValueError(f"{path}: row {row_number}: driver and object must not be empty")
        value = _cell_decimal(record["value"], path=path, row=row_number, column="value")
        per_driver = raw.setdefault(driver, {})
        # Duplicate driver/object pairs (e.g. monthly figures) are summed,
        # matching the CSV reader.
        per_driver[obj] = per_driver.get(obj, Decimal(0)) + value
    if not raw:
        raise ValueError(f"{path}: no driver rows found")
    return {
        driver: Distribution.from_values(values, negatives=negatives)
        for driver, values in raw.items()
    }


def read_costs_xlsx(
    path: PathLike, *, sheet: Optional[str] = None
) -> list[CostLine]:
    """Read cost lines from an Excel sheet with columns ``cost,amount,driver``."""
    lines: list[CostLine] = []
    for row_number, record in _iter_rows(path, COST_COLUMNS, sheet):
        name = _cell_text(record["cost"])
        driver = _cell_text(record["driver"])
        if not name or not driver:
            raise ValueError(f"{path}: row {row_number}: cost and driver must not be empty")
        amount = _cell_decimal(record["amount"], path=path, row=row_number, column="amount")
        lines.append(CostLine(name=name, amount=amount, driver=driver))
    if not lines:
        raise ValueError(f"{path}: no cost rows found")
    return lines


def write_allocation_xlsx(
    path: PathLike, rows: list[AllocationRow], *, precision: int = 2
) -> None:
    """Write allocation rows to an Excel file, one row per cost x object."""
    openpyxl = _openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "allocation"
    worksheet.append(["cost", "driver", "object", "share_percent", "amount"])
    amount_format = "0." + "0" * precision if precision > 0 else "0"
    for row in rows:
        worksheet.append(
            [row.cost, row.driver, row.object, float(row.share * 100), float(row.amount)]
        )
        worksheet.cell(row=worksheet.max_row, column=4).number_format = "0.0000"
        worksheet.cell(row=worksheet.max_row, column=5).number_format = amount_format
    workbook.save(path)


def write_percentages_xlsx(
    path: PathLike, percentages: Mapping[str, Mapping[str, Decimal]]
) -> None:
    """Write per-driver percentages to an Excel file (``driver,object,percent``)."""
    openpyxl = _openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "percentages"
    worksheet.append(["driver", "object", "percent"])
    for driver in sorted(percentages):
        for obj in sorted(percentages[driver]):
            worksheet.append([driver, obj, float(percentages[driver][obj])])
    workbook.save(path)
