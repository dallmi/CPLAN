"""End to end: raw CSVs in, a seven-sheet workbook out."""

from datetime import date

import pytest

pytest.importorskip("openpyxl")
from openpyxl import load_workbook

import pipeline.scripts.report_calendar as report_calendar
from tests.report_fixtures import write_activity_csvs

EXPECTED_SHEETS = [
    "Executive Summary", "Calendar", "Data Quality",
    "Audience & leadership", "Mix & Lead Time", "Activities", "Glossary",
]


@pytest.fixture(autouse=True)
def _sealed_repo_dir(tmp_path, monkeypatch):
    """The default GEB list is whatever `default_path(REPO_DIR)` finds --
    ambient state that lives outside `tmp_path`, in either format. The README
    tells the user to keep `geb-members.xlsx` or `geb-members.csv` in the
    repository root, and a developer who has done that (or left a real list
    behind from
    an earlier manual run) would otherwise leak a real membership list into
    every test exercising the default path -- including
    `test_without_a_list_the_workbook_is_unchanged`, the regression that
    matters most, which would then silently stop testing the absent-list case
    its name promises. Pointing REPO_DIR at this test's own `tmp_path` keeps
    the default path sealed inside the sandbox no matter what sits in the
    real repository root.
    """
    monkeypatch.setattr(report_calendar, "REPO_DIR", tmp_path)


def test_the_script_writes_all_seven_sheets(tmp_path):
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all", "--out", str(out)])

    assert code == 0
    assert out.exists()
    assert load_workbook(out).sheetnames == EXPECTED_SHEETS


def test_the_workbook_reopens_without_repair(tmp_path):
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"
    report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all", "--out", str(out)])

    wb = load_workbook(out)

    # Reading the calendar back proves the outline and merged-cell XML is valid.
    assert wb["Calendar"].cell(row=1, column=1).value == "Scope / activity"


def test_an_empty_input_directory_fails_loudly(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()

    code = report_calendar.main(["--input-dir", str(empty), "--out", str(tmp_path / "x.xlsx")])

    assert code == 1


def test_the_default_run_covers_every_dated_activity(tmp_path):
    """The fixtures carry a 2024 row that the old hard-coded 2025 window cut."""
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"
    report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all", "--out", str(out)])

    names = [row[1].value for row in load_workbook(out)["Activities"].iter_rows()]

    assert "Outside the window" in names
    assert dict_of_summary(out)["Period"] == "all dates"


def test_the_shipped_workbook_carries_the_geb_block_and_names(tmp_path):
    """End to end, against the CONFIG the script actually ships.

    The sheet-level tests build their own ReportConfig and so only prove the
    dataclass default. This one caught the block missing from a real run because
    the CONFIG block overrode `breakdown_fields` with its own pair.
    """
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"
    report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all", "--out", str(out)])
    wb = load_workbook(out)

    calendar = wb["Calendar"]
    labels = [calendar.cell(row=r, column=1).value
              for r in range(1, calendar.max_row + 1)]
    assert "BY GEB/GEB-1 — multiple values possible" in labels

    activities = wb["Activities"]
    headers = [activities.cell(row=1, column=c).value
               for c in range(1, activities.max_column + 1)]
    members = headers.index("GEB/GEB-1 members") + 1
    named = [activities.cell(row=r, column=members).value
             for r in range(2, activities.max_row + 1)]
    assert any(name for name in named), "no activity carries a GEB/GEB-1 member"


def test_the_shipped_config_actually_excludes_the_catch_all_objective():
    """Against the CONFIG the script ships, not a test-built one.

    The fixtures carry no catch-all row, so an end-to-end run cannot show this
    filter working; and a CONFIG that quietly lost the prefix would still
    produce a perfectly valid workbook. Same gap that hid the missing calendar
    block, so it gets the same kind of test.
    """
    import pandas as pd

    from pipeline.report.data import build_scope
    from pipeline.scripts.process_cplan import ActivityLoad

    frame = pd.DataFrame({
        "tracking_id": ["A", "B", "C"],
        "activity_name": ["Catch-all only", "Catch-all plus real", "Real only"],
        "start_date": pd.to_datetime(["2026-03-05"] * 3),
        "strategic_objectives": ["2026: Other", "2026: Other, 2026: Growth",
                                 "2026: Growth"],
    })

    scope = build_scope(ActivityLoad(frame, {}, {}), report_calendar.CONFIG)

    assert sorted(scope.frame["tracking_id"]) == ["B", "C"]
    assert scope.excluded["objectives"] == 1


def test_a_year_run_cuts_the_activities_outside_it(tmp_path):
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"
    report_calendar.main(["--input-dir", str(tmp_path / "input"),
                          "--year", "2025", "--out", str(out)])

    names = [row[1].value for row in load_workbook(out)["Activities"].iter_rows()]

    assert "Outside the window" not in names
    assert dict_of_summary(out)["Period"] == "2025-01-01 to 2025-12-31"


def dict_of_summary(path):
    """Label -> value from the Executive Summary, whatever column they sit in."""
    labels = {}
    for row in load_workbook(path)["Executive Summary"].iter_rows():
        values = [cell.value for cell in row if cell.value is not None]
        if len(values) >= 2:
            labels[values[0]] = values[1]
    return labels


# --- the period on the command line -----------------------------------------

def _resolve(argv):
    parser = report_calendar.build_parser()
    return report_calendar.resolve_config(
        report_calendar.CONFIG, parser.parse_args(argv), parser)


def test_no_period_flags_leave_the_config_block_alone():
    config = _resolve([])

    assert config == report_calendar.CONFIG
    assert config.date_from == date(2026, 1, 1)
    assert config.date_to == date(2026, 12, 31)


def test_all_clears_the_period_the_config_block_sets():
    config = _resolve(["--all"])

    assert config.date_from is None and config.date_to is None
    # Only the period is cleared; the other criteria are untouched.
    assert config.exclude_priorities == report_calendar.CONFIG.exclude_priorities


def test_all_cannot_be_combined_with_a_window():
    for argv in (["--all", "--year", "2026"], ["--all", "--from", "2026-01-01"]):
        with pytest.raises(SystemExit):
            _resolve(argv)


def test_year_expands_to_the_whole_calendar_year():
    config = _resolve(["--year", "2026"])

    assert config.date_from == date(2026, 1, 1)
    assert config.date_to == date(2026, 12, 31)


def test_from_and_to_set_the_window_verbatim():
    config = _resolve(["--from", "2025-01-01", "--to", "2026-12-31"])

    assert config.date_from == date(2025, 1, 1)
    assert config.date_to == date(2026, 12, 31)
    assert config.period_slug() == "2025-2026"


def test_from_alone_leaves_the_upper_bound_open():
    config = _resolve(["--from", "2026-04-01"])

    assert config.date_from == date(2026, 4, 1)
    assert config.date_to is None


def test_the_other_criteria_survive_a_period_flag():
    config = _resolve(["--year", "2026"])

    assert config.breakdown_fields == report_calendar.CONFIG.breakdown_fields
    assert config.include_archived == report_calendar.CONFIG.include_archived


def test_year_together_with_from_is_refused_rather_than_silently_winning():
    with pytest.raises(SystemExit):
        _resolve(["--year", "2026", "--from", "2026-04-01"])


def test_a_reversed_window_is_refused():
    with pytest.raises(SystemExit):
        _resolve(["--from", "2026-12-31", "--to", "2026-01-01"])


def test_a_malformed_date_is_refused_at_parse_time():
    with pytest.raises(SystemExit):
        _resolve(["--from", "01.04.2026"])


def test_the_default_output_lands_in_the_reports_folder():
    """Nothing but .xlsx lives there. The rest of pipeline/output is the
    pipeline's own working data -- parquet the API reads, meta.json the
    dashboard fetches -- and a delivered workbook was easy to lose among them.
    """
    path = report_calendar.default_output_path(_resolve([]))

    assert path.parent == report_calendar.REPORTS_DIR
    assert path.parent.name == "reports"
    assert path.parent.parent == report_calendar.OUTPUT_DIR


def test_an_explicit_out_path_still_wins(tmp_path):
    """The folder is a default, not a cage: --out puts the file where asked."""
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "somewhere" / "else.xlsx"

    code = report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all",
                                 "--out", str(out)])

    assert code == 0
    assert out.exists()


@pytest.mark.parametrize("argv,expected", [
    (["--all"], "CPLAN_calendar_all_"),
    ([], "CPLAN_calendar_2026_"),
    (["--year", "2026"], "CPLAN_calendar_2026_"),
    (["--from", "2025-01-01", "--to", "2026-12-31"], "CPLAN_calendar_2025-2026_"),
    (["--from", "2026-04-01", "--to", "2026-09-30"], "CPLAN_calendar_2026-04-01-2026-09-30_"),
])
def test_the_default_output_path_names_the_period(argv, expected):
    path = report_calendar.default_output_path(_resolve(argv))

    assert path.name.startswith(expected)
    assert path.suffix == ".xlsx"


def test_a_control_character_in_the_source_does_not_kill_the_run(tmp_path):
    """A real export carried a vertical tab at the end of an activity title.
    openpyxl raises on those rather than escaping them, and the exception landed
    mid-write: the whole run died and left no workbook behind. One bad byte in
    one row must at worst cost that character.
    """
    input_dir = tmp_path / "input"
    write_activity_csvs(input_dir)
    csv = input_dir / "InternalCommunicationActivities.csv"
    csv.write_text(csv.read_text(encoding="utf-8")
                   .replace("Single division Q1", "Single division Q1\x0b"),
                   encoding="utf-8")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main(["--input-dir", str(input_dir), "--all", "--out", str(out)])

    assert code == 0
    assert out.exists()
    names = [row[1].value for row in load_workbook(out)["Activities"].iter_rows()]
    assert "Single division Q1" in names


def test_an_unbounded_run_warns_when_one_date_stretches_the_whole_grid(tmp_path, capsys):
    """A mistyped year is the usual cause of a calendar that spans decades, and
    without a warning it just looks like a slow run. Name the rows at the edges,
    which is where the typo is.
    """
    input_dir = tmp_path / "input"
    write_activity_csvs(input_dir)
    csv = input_dir / "InternalCommunicationActivities.csv"
    csv.write_text(csv.read_text(encoding="utf-8").replace("2025-11-04", "2049-11-04"),
                   encoding="utf-8")

    report_calendar.main(["--input-dir", str(input_dir), "--all", "--out", str(tmp_path / "r.xlsx")])
    out = capsys.readouterr().out

    assert "WARNING" in out
    assert "2049-11-04" in out, "the warning must name the outlier"
    assert "Region only" in out, "and the activity carrying it"


def test_a_normal_span_says_nothing(tmp_path, capsys):
    write_activity_csvs(tmp_path / "input")

    report_calendar.main(["--input-dir", str(tmp_path / "input"), "--all",
                          "--out", str(tmp_path / "r.xlsx")])

    # Narrowed to the wide-grid warning specifically -- not a blanket
    # "WARNING" absence, which would also (correctly) trip on an unrelated
    # GEB-list warning were one ever in play here.
    assert "years wide" not in capsys.readouterr().out


def test_a_bounded_run_never_warns_however_wide_it_is(tmp_path, capsys):
    """The warning is about an accident. A window someone typed is a decision."""
    write_activity_csvs(tmp_path / "input")

    report_calendar.main(["--input-dir", str(tmp_path / "input"),
                          "--from", "2000-01-01", "--to", "2049-12-31",
                          "--out", str(tmp_path / "r.xlsx")])

    assert "years wide" not in capsys.readouterr().out


# --- the GEB membership list on the command line ----------------------------

GEB_EXAMPLE = 'email,name\n,"Example, Ada"\n'


def test_without_a_list_the_workbook_is_unchanged(tmp_path):
    """The regression that matters most. Every machine without the file must
    get byte-comparable sheet names and the combined block titles.
    """
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"

    report_calendar.main(["--input-dir", str(tmp_path / "input"),
                          "--all", "--out", str(out)])
    wb = load_workbook(out)

    assert wb.sheetnames == EXPECTED_SHEETS
    labels = [wb["Calendar"].cell(row=r, column=1).value
              for r in range(1, wb["Calendar"].max_row + 1)]
    assert "BY GEB/GEB-1 — multiple values possible" in labels
    assert "BY GEB — multiple values possible" not in labels


def test_a_list_splits_the_calendar_end_to_end(tmp_path):
    write_activity_csvs(tmp_path / "input")
    members = tmp_path / "geb-members.csv"
    members.write_text(GEB_EXAMPLE, encoding="utf-8")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main([
        "--input-dir", str(tmp_path / "input"), "--all",
        "--geb-members", str(members), "--out", str(out)])

    assert code == 0
    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_SHEETS  # no new sheet, only new blocks
    labels = [wb["Calendar"].cell(row=r, column=1).value
              for r in range(1, wb["Calendar"].max_row + 1)]
    assert "BY GEB — multiple values possible" in labels
    assert "BY GEB-1 — multiple values possible" in labels
    assert "BY GEB/GEB-1 — multiple values possible" not in labels


def test_a_broken_list_aborts_with_a_message(tmp_path, capsys):
    """Aborting beats falling back: a fallback produces a workbook that looks
    correct and is wrong.
    """
    write_activity_csvs(tmp_path / "input")
    members = tmp_path / "geb-members.csv"
    members.write_text("email\nonly@example.invalid\n", encoding="utf-8")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main([
        "--input-dir", str(tmp_path / "input"), "--all",
        "--geb-members", str(members), "--out", str(out)])

    assert code == 1
    assert not out.exists()
    # The real message, not a substring a raw traceback could also contain --
    # "fieldnames" (a DictReader attribute that would appear in an unguarded
    # traceback) also contains "name".
    assert "missing the required column 'name'" in capsys.readouterr().out


def test_a_named_list_that_does_not_exist_aborts(tmp_path):
    """An absent default file is normal; an absent *named* file is a typo on
    the command line and must not silently produce the unsplit workbook.
    """
    write_activity_csvs(tmp_path / "input")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main([
        "--input-dir", str(tmp_path / "input"), "--all",
        "--geb-members", str(tmp_path / "nope.csv"), "--out", str(out)])

    assert code == 1
    assert not out.exists()


def test_the_unmatched_count_is_logged_as_a_warning(tmp_path, capsys):
    """The console warning at report_calendar.py's unmatched-count check --
    the operator running the report is the person who can fix the list, and
    that only works if the console actually says so.
    """
    write_activity_csvs(tmp_path / "input")
    members = tmp_path / "geb-members.csv"
    members.write_text(
        'email,name\n,"Example, Ada"\n,"Nobody, Really"\n', encoding="utf-8")
    out = tmp_path / "report.xlsx"

    code = report_calendar.main([
        "--input-dir", str(tmp_path / "input"), "--all",
        "--geb-members", str(members), "--out", str(out)])

    assert code == 0
    log = capsys.readouterr().out
    assert "WARNING: 1 of 2 GEB list entries matched nothing in scope" in log
