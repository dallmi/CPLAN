"""Data Quality turns the pack problem into a figure."""

from datetime import date

import pandas as pd
import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook

from pipeline.report.config import ReportConfig
from pipeline.report.data import build_scope
from pipeline.report.table_sheets import build_data_quality
from pipeline.scripts.process_cplan import ActivityLoad
from tests.report_fixtures import load_fixture_scope


def _sheet(tmp_path):
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config)
    wb = Workbook()
    wb.remove(wb.active)
    build_data_quality(wb, scope, config)
    return wb["Data Quality"], scope


def _column_a(ws):
    return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]


def test_the_sheet_has_the_three_blocks(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)

    assert "FIELD COMPLETENESS" in labels
    assert "PACK COVERAGE" in labels
    assert "RECORD ANOMALIES" in labels


def test_field_rows_carry_a_missing_share_formula(tmp_path):
    ws, _ = _sheet(tmp_path)
    formulas = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row + 1)]

    assert any(isinstance(v, str) and v.startswith("=IF(") for v in formulas)


def test_pack_coverage_names_the_oversized_bucket(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = [str(v) for v in _column_a(ws)]

    assert any("more than 50" in label for label in labels)
    assert any("exactly one" in label for label in labels)


def test_the_counts_add_up_to_the_scope(tmp_path):
    ws, scope = _sheet(tmp_path)
    rows = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    with_pack = ws.cell(row=rows["Activities with a pack link"], column=2).value
    without_pack = ws.cell(row=rows["Activities without a pack link"], column=2).value

    assert with_pack + without_pack == len(scope.frame)


def test_the_anomaly_labels_match_what_the_figures_measure(tmp_path):
    """Task 9 renamed these two rows; Task 11 must not resurrect the old names."""
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)

    assert "Duplicate tracking IDs removed on load" in labels
    assert "Blank tracking ID (after de-duplication)" in labels


def test_duplicate_count_reads_the_real_figure_not_a_post_dedup_zero(tmp_path):
    """`anomalies()` takes `duplicates_removed` as a second argument now.

    The fixture carries one genuine cross-list duplicate (a stale archived
    copy of IC-0001). `load_activities` removes it before `build_scope` ever
    sees the frame, so a duplicate count computed only from the in-scope
    frame is structurally always zero. The builder must pass
    `scope.duplicates_removed` through, or this regresses to 0.
    """
    ws, scope = _sheet(tmp_path)
    rows = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}

    assert scope.duplicates_removed == 1
    dup_row = rows["Duplicate tracking IDs removed on load"]
    assert ws.cell(row=dup_row, column=2).value == 1


def test_median_completeness_has_its_own_block_that_names_its_denominator(tmp_path):
    """The per-field rates above are one row per *reported* field; the median
    is computed over the fields the entry form requires, split by source type.
    Both are correct and they are not the same denominator -- on the fixture
    the block shows bod_geb heavily missing and then a 100% median, which
    reads as a contradiction when the two sit in one block. The median gets
    its own labelled block, and this sheet names the fields it counts.
    """
    ws, scope = _sheet(tmp_path)
    labels = [str(v) for v in _column_a(ws)]

    assert "PLANNING COMPLETENESS" in labels
    assert "Median completeness (%)" not in labels  # the old, unqualified label
    assert labels.index("Median planning completeness (%)") > labels.index(
        "PLANNING COMPLETENESS")
    assert labels.index("PLANNING COMPLETENESS") > labels.index("FIELD COMPLETENESS")

    text = "\n".join(str(ws.cell(row=r, column=c).value)
                     for r in range(1, ws.max_row + 1) for c in (1, 2))
    assert scope.completeness_fields  # sanity: the fixture carries some
    for name in scope.completeness_fields:
        assert name in text, f"{name} is counted but not named on the sheet"
    for name in scope.skipped_completeness_fields:
        assert name in text, f"{name} is excluded but not named on the sheet"


def test_the_sheet_builds_when_the_export_carries_no_end_date_column(tmp_path):
    """A missing column has to cost one honest gap, not the whole workbook:
    the RECORD ANOMALIES block used to raise AttributeError here and take
    every other sheet down with it.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    frame = pd.DataFrame([{
        "tracking_id": "IC-0001", "activity_name": "A", "source_type": "internal",
        "start_date": pd.Timestamp("2025-03-05"), "channel": "Email",
    }])
    assert "end_date" not in frame.columns
    scope = build_scope(ActivityLoad(frame, {}, {}), config)

    wb = Workbook()
    wb.remove(wb.active)
    build_data_quality(wb, scope, config)
    ws = wb["Data Quality"]
    labels = [str(v) for v in _column_a(ws)]

    assert "RECORD ANOMALIES" in labels
    assert ws.cell(row=labels.index("Missing end date") + 1, column=2).value == 1


def test_an_empty_scope_does_not_crash(tmp_path):
    """Task 10 shipped a Critical bug on exactly this shape: a builder that
    raised on the frame `build_scope` produces when nothing was read at all
    -- no columns, not merely no rows. This sheet must degrade gracefully
    instead.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    load = ActivityLoad(pd.DataFrame(), {}, {})
    scope = build_scope(load, config)
    assert scope.frame.empty
    assert list(scope.frame.columns) == []

    wb = Workbook()
    wb.remove(wb.active)
    build_data_quality(wb, scope, config)
    ws = wb["Data Quality"]

    assert ws.cell(row=1, column=1).value is not None
