"""Audience bands and executive involvement, by period and by division."""

from datetime import date

import pandas as pd
import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pipeline.report.config import AUDIENCE_BANDS, BAND_UNKNOWN, ReportConfig
from pipeline.report.data import build_scope
from pipeline.report.table_sheets import build_audience
from pipeline.scripts.process_cplan import ActivityLoad
from tests.report_fixtures import load_fixture_scope


def _sheet(tmp_path):
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config)
    wb = Workbook()
    wb.remove(wb.active)
    build_audience(wb, scope, config)
    return wb["Audience & GEB"], scope


def _column_a(ws):
    return [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]


def test_every_band_gets_a_row_including_unknown(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)

    for band in AUDIENCE_BANDS:
        assert band in labels
    assert BAND_UNKNOWN in labels


def test_the_band_table_totals_are_formulas(tmp_path):
    ws, _ = _sheet(tmp_path)
    row = _column_a(ws).index(AUDIENCE_BANDS[0]) + 1
    last_column = ws.max_column

    assert str(ws.cell(row=row, column=last_column - 1).value).startswith("=SUM(")
    assert str(ws.cell(row=row, column=last_column).value).startswith("=IF(")


def test_the_sheet_reports_executives_by_quarter_and_by_division(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)

    assert "GEB INVOLVEMENT BY QUARTER" in labels
    assert "GEB INVOLVEMENT BY DIVISION" in labels


def test_division_rows_report_the_share_of_that_divisions_own_volume(tmp_path):
    ws, _ = _sheet(tmp_path)
    headers = [str(ws.cell(row=r, column=4).value) for r in range(1, ws.max_row + 1)]

    assert any("of the division" in header for header in headers)


def test_the_band_row_formulas_reference_the_cells_they_belong_to(tmp_path):
    """A formula pointing at the wrong cell still renders a plausible
    percentage. Check the actual references, not just the formula's shape.
    """
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)
    row = labels.index(AUDIENCE_BANDS[0]) + 1
    total_row = labels.index("TOTAL") + 1
    last_column = ws.max_column
    total_col = last_column - 1
    share_col = last_column
    total_letter = get_column_letter(total_col)
    quarter_span_end = get_column_letter(total_col - 1)

    total_formula = str(ws.cell(row=row, column=total_col).value)
    assert total_formula == f"=SUM(B{row}:{quarter_span_end}{row})"

    share_formula = str(ws.cell(row=row, column=share_col).value)
    assert share_formula == (
        f"=IF({total_letter}${total_row}=0,0,"
        f"{total_letter}{row}/{total_letter}${total_row})"
    )


def test_the_large_audience_by_month_block_divides_its_own_row(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _column_a(ws)
    assert "LARGE AUDIENCE BY MONTH" in labels
    header_row = labels.index("LARGE AUDIENCE BY MONTH") + 1  # section header row
    data_row = header_row + 2  # + column-header row + first data row
    formula = str(ws.cell(row=data_row, column=4).value)
    assert formula == f"=IF(C{data_row}=0,0,B{data_row}/C{data_row})"


# --- ACTIVITIES BY GEB MEMBER ------------------------------------------------

def _geb_sheet(sources):
    """A sheet built from nothing but start dates and the GEB source field."""
    frame = pd.DataFrame({
        "start_date": pd.to_datetime(["2025-03-05"] * len(sources)),
        "bod_geb": list(sources),
        "activity_name": [f"Activity {i}" for i in range(len(sources))],
    })
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = build_scope(ActivityLoad(frame, {}, {}), config)
    wb = Workbook()
    wb.remove(wb.active)
    build_audience(wb, scope, config)
    return wb["Audience & GEB"]


def _member_block(ws, title="ACTIVITIES BY GEB MEMBER"):
    """Label -> count for the person rows under one block.

    Stops at the next section header: a second people block follows this one,
    and a helper that ran to max_row would silently merge the two.
    """
    labels = _column_a(ws)
    start = labels.index(title) + 3  # + column-header row + denominator row
    out = {}
    for r in range(start + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        if str(label).isupper() and str(label).startswith("ACTIVITIES BY"):
            break
        out[label] = ws.cell(row=r, column=2).value
    return out


def test_each_named_member_gets_a_row_with_their_own_count():
    ws = _geb_sheet(["A. Person", "A. Person", "B. Person", ""])

    assert _member_block(ws) == {"A. Person": 2, "B. Person": 1}


def test_an_activity_naming_two_members_counts_under_both():
    ws = _geb_sheet(["A. Person; B. Person", "A. Person"])

    assert _member_block(ws) == {"A. Person": 2, "B. Person": 1}


def test_the_denominator_is_the_distinct_count_of_involved_activities():
    """Two members on one activity is still one activity with GEB."""
    ws = _geb_sheet(["A. Person; B. Person", "", ""])
    labels = _column_a(ws)
    denominator_row = labels.index("ACTIVITIES BY GEB MEMBER") + 3

    assert ws.cell(row=denominator_row, column=1).value == "All activities with GEB"
    assert ws.cell(row=denominator_row, column=2).value == 1


def test_each_share_divides_by_that_denominator_cell_rather_than_a_literal():
    ws = _geb_sheet(["A. Person", "B. Person"])
    labels = _column_a(ws)
    denominator_row = labels.index("ACTIVITIES BY GEB MEMBER") + 3
    first_member = denominator_row + 1

    formula = str(ws.cell(row=first_member, column=3).value)

    assert formula == (f"=IF($B${denominator_row}=0,0,"
                       f"B{first_member}/$B${denominator_row})")


def test_members_are_ordered_by_volume_then_by_name():
    ws = _geb_sheet(["B. Person", "B. Person", "C. Person", "A. Person"])

    assert list(_member_block(ws)) == ["B. Person", "A. Person", "C. Person"]


def test_a_scope_without_any_named_member_says_so_instead_of_showing_nothing():
    ws = _geb_sheet(["", "   "])

    assert "No GEB member named on any in-scope activity" in _column_a(ws)


def test_an_empty_scope_does_not_crash():
    """Task 10 shipped a Critical bug on exactly this shape: a builder that
    raised on the frame `build_scope` produces when nothing was read at all
    -- no columns, not merely no rows. This sheet must degrade gracefully
    instead.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    load = ActivityLoad(pd.DataFrame(), {}, {})
    scope = build_scope(load, config)
    assert scope.frame.empty
    assert list(scope.frame.columns) == []

    wb = Workbook()
    wb.remove(wb.active)
    build_audience(wb, scope, config)
    ws = wb["Audience & GEB"]

    assert ws.cell(row=1, column=1).value is not None
