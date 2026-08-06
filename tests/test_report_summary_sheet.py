"""The summary carries the criteria, the volume, the load and the caveats."""

from datetime import date

import pandas as pd
import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook

from pipeline.report.config import AUDIENCE_BAND_ORDER, ReportConfig
from pipeline.report.data import build_scope
from pipeline.report.table_sheets import build_executive_summary, build_glossary
from pipeline.scripts.process_cplan import ActivityLoad
from tests.report_fixtures import load_fixture_scope


def _members(*names):
    from pipeline.report.membership import Entry, Membership, normalise_name
    return Membership(entries=tuple(
        Entry(email="", name=normalise_name(n)) for n in names))


def _build(tmp_path, builder):
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config)
    wb = Workbook()
    wb.remove(wb.active)
    builder(wb, scope, config)
    return wb.worksheets[0], scope


def _build_with_membership(tmp_path, builder, membership):
    """Like `_build`, but with a membership list loaded.

    Every pre-existing sheet-level test built through `_build` alone, which
    never exercises the membership branch -- exactly the gap that let the
    GEB share row's denominator, and the Glossary's split terms, go
    unmeasured. New assertions about membership-only behaviour should build
    through here instead.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config, membership=membership)
    wb = Workbook()
    wb.remove(wb.active)
    builder(wb, scope, config)
    return wb.worksheets[0], scope


def _pairs(ws):
    return {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)}


def test_the_summary_states_the_applied_criteria(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    pairs = _pairs(ws)

    assert pairs["Period"] == "2025-01-01 to 2025-12-31"
    assert pairs["GEB/GEB-1"] == "any"


def test_the_summary_shows_readable_breakdown_dimension_names(tmp_path):
    """`describe()` used to print the raw field identifiers verbatim; a reader
    would see "executives" rather than "GEB/GEB-1" in the REPORT section.
    """
    ws, _ = _build(tmp_path, build_executive_summary)
    pairs = _pairs(ws)

    assert pairs["Breakdown dimensions"] == "BUSINESS DIVISION, REGION, COUNTRY, GEB/GEB-1"


def test_the_summary_shows_readable_breakdown_dimension_names_with_a_membership(tmp_path):
    """Once a membership splits `executives` into `executives_geb` and
    `executives_geb1`, `report_calendar` swaps those into `breakdown_fields`
    -- the raw identifiers must still render as GEB, GEB-1, not themselves.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31),
                          breakdown_fields=("business_division", "region_group", "country",
                                            "executives_geb", "executives_geb1"))
    scope = load_fixture_scope(tmp_path, config, membership=_members("Example, Ada"))
    wb = Workbook()
    wb.remove(wb.active)
    build_executive_summary(wb, scope, config)
    pairs = _pairs(wb.worksheets[0])

    assert pairs["Breakdown dimensions"] == "BUSINESS DIVISION, REGION, COUNTRY, GEB, GEB-1"


def test_the_summary_does_not_name_the_operator_s_source_files(tmp_path):
    """Removed 2026-08-06 on request: the rows confused recipients.

    The workbook is forwarded to people who have never seen the directory it
    was built from, so an export filename there answers nobody's question and
    raises several -- and filenames carry dates and initials the sheet then
    cannot explain. The provenance is not lost: `report_calendar.py` logs it
    to the console, where the person who chose the files is the one reading.
    """
    ws, scope = _build(tmp_path, build_executive_summary)
    cells = [str(ws.cell(row=r, column=c).value)
             for r in range(1, ws.max_row + 1) for c in range(1, 3)]

    assert scope.source_files, "the fixture must actually have source files"
    for _, name in scope.source_files:
        assert not any(name in cell for cell in cells), \
            f"the Executive Summary still names the source file {name!r}"
    assert not any(cell.startswith("Source") for cell in cells)


def test_the_summary_reports_what_each_criterion_excluded(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    pairs = _pairs(ws)

    assert pairs["Excluded: no start date"] == 1
    assert pairs["Excluded: date window"] == 1


def test_shares_are_formulas_not_baked_numbers(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    labels = [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]

    assert any(label.startswith("=TEXT(IF(") for label in labels)


def test_the_summary_still_renders_the_report_section_on_an_empty_scope():
    """Nothing in scope is exactly when the REPORT section -- the criteria,
    the source files, which filter removed what -- matters most. It must
    render rather than raise, even though the empty-load path never attaches
    the derived columns (week_index included) that the other sections read.
    """
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = build_scope(ActivityLoad(pd.DataFrame(), {}, {}), config)
    wb = Workbook()
    wb.remove(wb.active)

    build_executive_summary(wb, scope, config)

    ws = wb.worksheets[0]
    pairs = _pairs(ws)
    assert pairs["Period"] == "2025-01-01 to 2025-12-31"
    assert pairs["GEB/GEB-1"] == "any"
    assert pairs["Excluded: no start date"] == 0
    assert pairs["Excluded: date window"] == 0
    assert pairs["Rows read"] == 0


def _assert_every_share_divides_by_the_section_total(ws):
    """A formula that names the wrong denominator row still renders a
    plausible-looking percentage -- startswith("=TEXT(IF(") alone cannot
    catch that. Check the actual cell references instead.

    Runs over every `=TEXT(IF(` row on the sheet, whatever the sheet
    contains -- so a membership build's `With GEB involvement` row is
    checked the same way as every other share, without a separate branch.
    """
    rows = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    total_row = rows["Activities in scope"]

    checked = 0
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str) or not label.startswith("=TEXT(IF("):
            continue
        assert f"B${total_row}=0" in label, f"row {r} guards the wrong total"
        assert f"B{r}/B${total_row}" in label, f"row {r} divides the wrong cells"
        checked += 1
    return checked


def test_every_share_formula_divides_by_its_own_section_total(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    checked = _assert_every_share_divides_by_the_section_total(ws)
    assert checked >= 8  # internal/external plus the six audience bands


def test_the_geb_share_formula_also_divides_by_the_section_total(tmp_path):
    """The `With GEB involvement` row only exists on a membership build, so
    the check above alone never touches it. Left unpinned, a version of the
    row that divides by its neighbour instead (the combined `With
    GEB/GEB-1 involvement` row directly above it) still renders a plausible
    percentage on the fixture, and every other report test stays green.
    """
    ws, _ = _build_with_membership(tmp_path, build_executive_summary, _members("Example, Ada"))
    checked = _assert_every_share_divides_by_the_section_total(ws)
    assert checked >= 9  # the eight above, plus the new GEB row


def test_the_summary_reports_load_and_discipline(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    pairs = _pairs(ws)

    assert "Peak week" in pairs
    assert "Weeks with no activity" in pairs
    assert "Median lead time (days)" in pairs



def _glossary_entries(ws):
    """(term, definition) for every defined row, skipping the section bands."""
    out = []
    for r in range(1, ws.max_row + 1):
        term, definition = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if term and definition:
            out.append((str(term), str(definition)))
    return out


MAX_DEFINITION_CHARS = 110


def test_every_glossary_definition_stays_short(tmp_path):
    """Plain and short is the requirement, not a style preference.

    The Glossary drifted into paragraph-long justifications -- one entry ran to
    eight lines explaining why two completeness figures could look inconsistent.
    A reader who needs that much prose to understand a column has been failed by
    the column, not by the Glossary. This pins the ceiling so the drift cannot
    quietly happen again.
    """
    ws, _ = _build(tmp_path, build_glossary)
    entries = _glossary_entries(ws)

    assert entries, "the Glossary has no definitions at all"
    too_long = [(t, len(d)) for t, d in entries if len(d) > MAX_DEFINITION_CHARS]
    assert not too_long, f"definitions over {MAX_DEFINITION_CHARS} chars: {too_long}"

    # GEB_SPLIT_TERMS only renders on a membership build (see
    # `_glossary_sections`), so the no-membership build above never measures
    # it. Checked here rather than only for presence, so the ceiling actually
    # bounds every definition the workbook can print, not just the ones a
    # no-membership run happens to include -- and the positive case (the
    # terms are actually defined once a list is loaded, not merely short if
    # they were) is pinned in the same build.
    ws_with_list, _ = _build_with_membership(tmp_path, build_glossary, _members("Example, Ada"))
    entries_with_list = _glossary_entries(ws_with_list)
    terms_with_list = {t: d for t, d in entries_with_list}
    assert "GEB" in terms_with_list, "the Glossary omits GEB when a list is loaded"
    # GEB-1 is deliberately undefined since 2026-08-06 (see GEB_SPLIT_TERMS),
    # and so is the "everyone else in the field is GEB-1" clause GEB carried.
    # The sheets still print a GEB-1 heading; leaving it undefined was the
    # request. Restoring either is a product decision, not a regression fix.
    assert "GEB-1" not in terms_with_list
    assert "GEB-1" not in terms_with_list["GEB"]
    too_long_with_list = [(t, len(d)) for t, d in entries_with_list
                          if len(d) > MAX_DEFINITION_CHARS]
    assert not too_long_with_list, (
        f"definitions over {MAX_DEFINITION_CHARS} chars: {too_long_with_list}")


def test_the_glossary_defines_the_terms_a_reader_meets_on_the_sheets(tmp_path):
    """The terms that appear as column headers or row labels elsewhere.

    Deliberately absent, each removed on request: the data source, the note that
    studio-only activities never reach this report, the Thursday week-to-month
    rule, the comma/semicolon splitting caveat, and "Quarter delta" -- the Mix
    sheet no longer prints that column, and a glossary that defines a term the
    workbook never shows sends a reader hunting for it. They are recorded in the
    design document instead. Do not restore them here as a phantom regression --
    if they are wanted back, that is a product decision.
    """
    ws, _ = _build(tmp_path, build_glossary)
    terms = {term for term, _ in _glossary_entries(ws)}
    text = "\n".join(f"{t} {d}" for t, d in _glossary_entries(ws)).lower()

    for term in ("In scope", "Overlap", "Audience band",
                 "GEB/GEB-1", "Lead time", "Planning completeness",
                 "Weekly counts", "Packs"):
        assert term in terms, f"the Glossary does not define {term!r}"

    assert "Quarter delta" not in terms
    assert "delta" not in text

    assert "thursday" not in text
    assert "studio" not in text
    assert "semicolon" not in text
    assert ws.sheet_view.showGridLines is False


def test_the_export_now_carries_every_field_completeness_is_scored_against(tmp_path):
    """`time_zone` used to be the standing gap: present in the source, unmapped
    by the ETL, so every activity read as missing a time zone and the field had
    to be dropped from the denominator. It is mapped now, and nothing is skipped.
    """
    _, scope = _build(tmp_path, build_glossary)

    assert scope.skipped_completeness_fields == []
    assert "time_zone" in scope.completeness_fields


def test_the_glossary_lists_fields_the_export_does_not_carry():
    """The block still has to appear when a field really is absent -- an export
    that predates a form change is a real shape, not a hypothetical one.
    """
    frame = pd.DataFrame([{
        "tracking_id": "IC-0001", "activity_name": "A", "source_type": "internal",
        "start_date": pd.Timestamp("2025-03-05"), "channel": "Email",
    }])
    assert "time_zone" not in frame.columns

    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = build_scope(ActivityLoad(frame, {}, {}), config)
    wb = Workbook()
    wb.remove(wb.active)
    build_glossary(wb, scope, config)
    ws = wb.worksheets[0]
    column_a = [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]

    assert "time_zone" in scope.skipped_completeness_fields
    assert "FIELDS NOT IN THIS EXPORT" in column_a
    for name in scope.skipped_completeness_fields:
        assert name in column_a


# --- shares ride in the label, so column C is never one lonely percentage ----

def test_the_leadership_figures_carry_their_share_in_the_label(tmp_path):
    """The GEB/GEB-1 count used to sit as a bare number with a lone percentage in
    column C beside it, next to two rows that had none. Both figures now use
    the same TEXT() label formula the VOLUME breakdown uses.
    """
    ws, _ = _build(tmp_path, build_executive_summary)
    labels = [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]

    geb = [label for label in labels if "With GEB/GEB-1 involvement" in label]
    large = [label for label in labels if "Large audience" in label]

    assert geb and large
    for label in geb + large:
        assert label.startswith('=TEXT(IF(B$'), label
        assert '"0%"' in label


def test_no_stray_percentage_is_left_in_column_c(tmp_path):
    """Column C carried exactly one value: the orphan this replaced."""
    ws, _ = _build(tmp_path, build_executive_summary)

    column_c = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]

    assert not [value for value in column_c if value is not None]


def test_the_volume_block_breaks_the_portfolio_down_by_audience_band(tmp_path):
    """Reach buckets were replaced by the audience bands, which is what a
    planner actually acts on. Every band gets a row, Unknown included, because
    the block is a partition of the portfolio.
    """
    ws, _ = _build(tmp_path, build_executive_summary)
    labels = " ".join(str(ws.cell(row=r, column=1).value)
                      for r in range(1, ws.max_row + 1))

    for band in AUDIENCE_BAND_ORDER:
        assert band in labels, f"the VOLUME block does not list {band!r}"
    assert "Group-wide" not in labels
    assert "Single division" not in labels


# --- the GEB share, printed only once a membership list makes it legible ----

def test_the_summary_adds_a_geb_share_when_a_list_is_present(tmp_path):
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config, membership=_members("Example, Ada"))
    wb = Workbook()
    wb.remove(wb.active)
    build_executive_summary(wb, scope, config)
    labels = [str(wb.worksheets[0].cell(row=r, column=1).value)
              for r in range(1, wb.worksheets[0].max_row + 1)]

    assert any("With GEB involvement" in label for label in labels)
    assert any("With GEB/GEB-1 involvement" in label for label in labels)


def test_the_summary_omits_the_geb_share_without_a_list(tmp_path):
    ws, _ = _build(tmp_path, build_executive_summary)
    labels = [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]

    assert not any("With GEB involvement" in label for label in labels)


def test_the_glossary_defines_both_levels_only_with_a_list(tmp_path):
    ws, _ = _build(tmp_path, build_glossary)
    terms = {term for term, _ in _glossary_entries(ws)}

    assert "GEB/GEB-1" in terms
    assert "GEB" not in terms
    assert "GEB-1" not in terms
