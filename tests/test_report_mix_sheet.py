"""Mix over time, lead time by division, and the traceable detail list."""

from datetime import date

import pandas as pd
import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pipeline.report.config import ReportConfig
from pipeline.report.data import build_scope
from pipeline.report.derive import priority_rank
from pipeline.report.table_sheets import (
    ACTIVITY_COLUMNS,
    _quarter_label,
    build_activities,
    build_mix,
)
from pipeline.scripts.process_cplan import ActivityLoad
from tests.report_fixtures import load_fixture_scope


def _build(tmp_path, builder, name):
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    scope = load_fixture_scope(tmp_path, config)
    wb = Workbook()
    wb.remove(wb.active)
    builder(wb, scope, config)
    return wb[name], scope


def _column_a(ws):
    return [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]


def test_mix_covers_channel_priority_and_source_type(tmp_path):
    ws, _ = _build(tmp_path, build_mix, "Mix & Lead Time")
    labels = _column_a(ws)

    assert "CHANNEL BY QUARTER" in labels
    assert "PRIORITY BY QUARTER" in labels
    assert "INTERNAL VS EXTERNAL BY QUARTER" in labels
    assert "LEAD TIME BY DIVISION" in labels


def _header_rows(ws):
    """Every crosstab header row, as a list of its non-empty header cells."""
    rows = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value != "Value":
            continue
        headers = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        while headers and headers[-1] is None:
            headers.pop()
        rows.append((r, headers))
    return rows


def test_the_crosstabs_carry_no_quarter_to_quarter_delta_column(tmp_path):
    """Removed 2026-08-06 on request; this is the guard against it returning.

    It compared the first quarter in scope against the last full one -- a
    settled quarter against one still being filled in. Every row therefore read
    strongly negative, and what the column actually measured was how far ahead
    the planning had got, not how the mix had moved. Nothing in the column
    itself lets a reader tell those two readings apart, which is what made it
    worse than absent. Bringing it back is a product decision, not a fix.
    """
    ws, _ = _build(tmp_path, build_mix, "Mix & Lead Time")
    cells = [str(ws.cell(row=r, column=c).value)
             for r in range(1, ws.max_row + 1)
             for c in range(1, ws.max_column + 1)]

    assert not any("Δ" in cell for cell in cells)
    # U+2212, the minus sign the header joined its two quarter labels with. The
    # em dash the "no data" and "missing" cells use is a different character,
    # so this does not catch those.
    assert not any("−" in cell for cell in cells)

    rows = _header_rows(ws)
    assert len(rows) == 3, "expected the channel, priority and source-type blocks"
    for _, headers in rows:
        assert headers[-1] == "Total", f"Total is not the last column: {headers}"


def test_the_total_sums_exactly_the_quarter_columns(tmp_path):
    """Dropping the Δ column moved Total one place right. An off-by-one here
    prints a total that silently includes or omits a quarter, and nothing on
    the sheet looks wrong -- the number just does not add up.
    """
    ws, scope = _build(tmp_path, build_mix, "Mix & Lead Time")
    quarters = sorted({q for q in scope.frame["_quarter"] if q is not None})

    header_row, headers = _header_rows(ws)[0]
    assert headers == ["Value"] + [_quarter_label(q) for q in quarters] + ["Total"]

    first_col = get_column_letter(2)
    last_col = get_column_letter(len(quarters) + 1)
    value_row = header_row + 1
    assert ws.cell(row=value_row, column=len(headers)).value == \
        f"=SUM({first_col}{value_row}:{last_col}{value_row})"


def test_the_priority_block_is_ordered_by_rank_not_alphabetically(tmp_path):
    """Two priority vocabularies are live at once, so alphabetical order is
    actively wrong: it interleaves "3 - low priority" with "High" and puts
    Low above Medium. The fixture carries both vocabularies for exactly this.
    """
    ws, _ = _build(tmp_path, build_mix, "Mix & Lead Time")
    column_a = _column_a(ws)
    start = column_a.index("PRIORITY BY QUARTER")
    rows = []
    for label in column_a[start + 2:]:
        if label in ("None", "") or label.isupper():
            break
        rows.append(label)

    assert len(rows) > 1, "the fixture must carry more than one priority value"
    ranks = [priority_rank(label) for label in rows]
    assert ranks == sorted(ranks, reverse=True), f"not ranked most-urgent-first: {rows}"
    assert rows != sorted(rows), "alphabetical order would have passed this test"


def test_the_activities_sheet_lists_every_in_scope_row(tmp_path):
    ws, scope = _build(tmp_path, build_activities, "Activities")

    assert ws.max_row == len(scope.frame) + 1
    assert ws.auto_filter.ref is not None
    assert ws.freeze_panes == "A2"


def test_the_activities_sheet_carries_the_derived_columns(tmp_path):
    ws, _ = _build(tmp_path, build_activities, "Activities")
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    assert headers == [header for _, header in ACTIVITY_COLUMNS]
    assert "Audience band" in headers
    assert "GEB/GEB-1 involved" in headers
    assert "GEB/GEB-1 members" in headers


def _empty_scope():
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    load = ActivityLoad(pd.DataFrame(), {}, {})
    return build_scope(load, config)


def test_an_empty_scope_does_not_crash_the_mix_sheet():
    """Task 10 shipped a Critical bug on exactly this shape: a builder that
    raised on the frame `build_scope` produces when nothing was read at all
    -- no columns, not merely no rows. This sheet must degrade gracefully
    instead.
    """
    scope = _empty_scope()
    assert scope.frame.empty
    assert list(scope.frame.columns) == []

    wb = Workbook()
    wb.remove(wb.active)
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    build_mix(wb, scope, config)
    ws = wb["Mix & Lead Time"]

    assert ws.cell(row=1, column=1).value is not None


def test_an_empty_scope_does_not_crash_the_activities_sheet():
    scope = _empty_scope()
    assert scope.frame.empty
    assert list(scope.frame.columns) == []

    wb = Workbook()
    wb.remove(wb.active)
    config = ReportConfig(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    build_activities(wb, scope, config)
    ws = wb["Activities"]

    # No rows beyond the header -- the row count still equals the in-scope
    # count exactly (zero), it just never gets the chance to drop anything.
    assert ws.max_row == len(scope.frame) + 1
    assert [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)] == [
        header for _, header in ACTIVITY_COLUMNS
    ]


def test_the_activities_row_count_matches_the_in_scope_count_exactly(tmp_path):
    """The Activities sheet must not silently drop rows: this is the audit
    trail every other sheet's figures have to be traceable back to.
    """
    ws, scope = _build(tmp_path, build_activities, "Activities")
    assert len(scope.frame) > 0  # sanity: the fixture is not itself empty
    assert ws.max_row == len(scope.frame) + 1
