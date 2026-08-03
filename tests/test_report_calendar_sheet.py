"""The calendar matrix: outline levels, formulas, and the double-count trap."""

import re
from datetime import date

import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from pipeline.report.calendar_sheet import (
    FIRST_GRID_COL,
    LABEL_COL,
    TOTAL_COL,
    _children,
    build_calendar,
)
from pipeline.report.config import ReportConfig
from tests.report_fixtures import load_fixture_scope


def _config(**overrides):
    base = dict(date_from=date(2025, 1, 1), date_to=date(2025, 12, 31))
    base.update(overrides)
    return ReportConfig(**base)


def _sheet(tmp_path, **overrides):
    scope = load_fixture_scope(tmp_path, _config(**overrides))
    wb = Workbook()
    wb.remove(wb.active)
    build_calendar(wb, scope, _config(**overrides))
    return wb["Calendar"], scope


def _labels(ws):
    return {
        ws.cell(row=r, column=LABEL_COL).value: r
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=LABEL_COL).value
    }


def test_the_header_names_the_axes(tmp_path):
    ws, _ = _sheet(tmp_path)

    assert ws.cell(row=1, column=LABEL_COL).value == "Scope / activity"
    assert ws.cell(row=1, column=TOTAL_COL).value == "Total"
    assert ws.cell(row=1, column=FIRST_GRID_COL).value == "Q1 2025"
    assert ws.freeze_panes == "C3"


def test_columns_carry_the_three_outline_levels_and_open_collapsed(tmp_path):
    ws, _ = _sheet(tmp_path)
    levels = {}
    for letter, dimension in ws.column_dimensions.items():
        levels.setdefault(dimension.outline_level, []).append(letter)

    assert set(levels) >= {0, 1, 2}
    month_letter = levels[1][0]
    week_letter = levels[2][0]
    assert ws.column_dimensions[month_letter].hidden is True
    assert ws.column_dimensions[week_letter].hidden is True
    assert ws.sheet_properties.outlinePr.summaryRight is False


def test_rows_carry_block_dimension_and_activity_levels(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _labels(ws)
    block_row = labels["BY AUDIENCE"]

    assert ws.row_dimensions[block_row].outline_level == 0
    child_rows = [r for r in range(block_row + 1, ws.max_row + 1)
                  if ws.row_dimensions[r].outline_level == 1]
    assert child_rows
    assert ws.row_dimensions[child_rows[0]].hidden is True


def test_month_and_quarter_cells_are_sum_formulas_over_their_children(tmp_path):
    ws, _ = _sheet(tmp_path)
    row = _labels(ws)["ALL ACTIVITIES"]

    quarter_cell = ws.cell(row=row, column=FIRST_GRID_COL).value
    assert isinstance(quarter_cell, str) and quarter_cell.startswith("=SUM(")
    total_cell = ws.cell(row=row, column=TOTAL_COL).value
    assert isinstance(total_cell, str) and total_cell.startswith("=SUM(")


def test_week_cells_are_literal_counts(tmp_path):
    ws, scope = _sheet(tmp_path)
    row = _labels(ws)["ALL ACTIVITIES"]
    week_columns = [c for c in range(FIRST_GRID_COL, ws.max_column + 1)
                    if ws.column_dimensions[
                        ws.cell(row=1, column=c).column_letter].outline_level == 2]
    values = [ws.cell(row=row, column=c).value for c in week_columns]
    numeric = [v for v in values if isinstance(v, int)]

    assert sum(numeric) == len(scope.frame)


def test_the_audience_block_header_sums_its_children(tmp_path):
    ws, _ = _sheet(tmp_path)
    row = _labels(ws)["BY AUDIENCE"]

    assert str(ws.cell(row=row, column=TOTAL_COL).value).startswith("=SUM(")


def test_the_division_block_header_is_a_distinct_count_not_a_sum(tmp_path):
    """The literal is derived from the header row's own week counts, not from
    `len(scope.frame)` -- that is the property the implementation establishes,
    and asserting the frame length instead would still pass if the header ever
    stopped agreeing with the cells printed beside it.
    """
    ws, _ = _sheet(tmp_path)
    labels = _labels(ws)
    header = next(label for label in labels if label.startswith("BY BUSINESS DIVISION"))
    row = labels[header]

    week_columns = [c for c in range(FIRST_GRID_COL, ws.max_column + 1)
                    if ws.column_dimensions[
                        ws.cell(row=1, column=c).column_letter].outline_level == 2]
    own_weeks = sum(v for v in (ws.cell(row=row, column=c).value for c in week_columns)
                    if isinstance(v, int))

    assert "multiple values possible" in header
    assert isinstance(ws.cell(row=row, column=TOTAL_COL).value, str) is False
    assert ws.cell(row=row, column=TOTAL_COL).value == own_weeks


def _refs(formula):
    """The cell references a formula names, as (column letter, row) pairs."""
    return [(letter, int(number))
            for letter, number in re.findall(r"([A-Z]+)(\d+)", str(formula))]


def _expected_children(columns):
    """Which weeks belong to which month, and which months to which quarter.

    Derived from ISO 8601 directly -- a week belongs to the month containing
    its Thursday -- rather than from `_children`, which is the thing under
    test. An expectation computed by calling the implementation would agree
    with any off-by-one the implementation happened to have.
    """
    month_weeks = {}
    quarter_months = {}
    for column in columns:
        if column.kind != "week":
            continue
        iso_year, iso_week = column.key
        thursday = date.fromisocalendar(iso_year, iso_week, 4)
        month = (thursday.year, thursday.month)
        quarter = (thursday.year, (thursday.month - 1) // 3 + 1)
        month_weeks.setdefault(month, []).append(column.key)
        months = quarter_months.setdefault(quarter, [])
        if month not in months:
            months.append(month)
    return month_weeks, quarter_months


def test_children_maps_weeks_to_months_by_the_thursday_rule(tmp_path):
    """`_children` is what every month and quarter formula is built from. An
    off-by-one there -- January summing February's weeks -- changes only which
    cells the formulas name, so the sheet still looks entirely well-formed.
    """
    _, scope = _sheet(tmp_path)
    columns = scope.grid.columns()

    assert _children(columns, scope.grid) == _expected_children(columns)


def test_each_aggregate_formula_sums_exactly_the_cells_it_should(tmp_path):
    """A `_children()` off-by-one would pass every other test on this sheet:
    the cells would still hold SUM formulas and the horizontal totals would
    still add up to something. The design's Testing section asks for the real
    property -- each quarter cell is the sum of its months and each month the
    sum of its weeks -- so check the actual references, not the shape.
    """
    ws, scope = _sheet(tmp_path)
    columns = scope.grid.columns()
    # Positions are the grid's own column order, recomputed here rather than
    # taken from `_column_positions`, for the same reason as above.
    positions = {(column.kind, column.key): FIRST_GRID_COL + offset
                 for offset, column in enumerate(columns)}
    month_weeks, quarter_months = _expected_children(columns)

    expected_children = {}
    for column in columns:
        col = positions[(column.kind, column.key)]
        if column.kind == "month":
            keys = [("week", key) for key in month_weeks[column.key]]
        elif column.kind == "quarter":
            keys = [("month", key) for key in quarter_months[column.key]]
        else:
            continue
        expected_children[col] = [get_column_letter(positions[key]) for key in keys]
    quarter_letters = [get_column_letter(positions[("quarter", column.key)])
                       for column in columns if column.kind == "quarter"]

    # Every quarter is genuinely made of months, and every month of weeks --
    # otherwise the loop below could pass vacuously.
    assert len(expected_children) >= 13 + 4
    assert all(children for children in expected_children.values())

    checked = 0
    for row in range(3, ws.max_row + 1):
        if not ws.cell(row=row, column=LABEL_COL).value:
            continue
        for col, children in expected_children.items():
            formula = ws.cell(row=row, column=col).value
            assert _refs(formula) == [(letter, row) for letter in children], (
                f"R{row}C{col} sums the wrong cells: {formula!r}")
            checked += 1

        total = ws.cell(row=row, column=TOTAL_COL).value
        if isinstance(total, int):
            continue  # a distinct-count block header: a literal by design
        refs = _refs(total)
        if all(letter == "B" for letter, _ in refs):
            continue  # the audience header's vertical audit SUM down its members
        assert refs == [(letter, row) for letter in quarter_letters], (
            f"the Total cell on row {row} sums the wrong cells: {total!r}")
    assert checked > 0


def test_every_row_keeps_week_cells_literal_and_aggregates_as_formulas(tmp_path):
    ws, _ = _sheet(tmp_path)
    levels = {c: ws.column_dimensions[ws.cell(row=1, column=c).column_letter].outline_level
              for c in range(FIRST_GRID_COL, ws.max_column + 1)}

    for r in range(3, ws.max_row + 1):
        if not ws.cell(row=r, column=LABEL_COL).value:
            continue
        for c, level in levels.items():
            value = ws.cell(row=r, column=c).value
            if level == 2:
                assert not isinstance(value, str), f"week cell R{r}C{c} is a formula"
            else:
                assert value is None or str(value).startswith("=SUM("), \
                    f"aggregation cell R{r}C{c} is not a SUM"


def test_detail_rows_can_be_switched_off(tmp_path):
    with_detail, _ = _sheet(tmp_path, detail_rows=True)
    without_detail, _ = _sheet(tmp_path, detail_rows=False)

    assert with_detail.max_row > without_detail.max_row
    levels = {without_detail.row_dimensions[r].outline_level
              for r in range(3, without_detail.max_row + 1)}
    assert 2 not in levels


def test_the_outline_declares_its_depth_and_marks_its_groups_collapsed(tmp_path):
    """Hidden alone is not an outline.

    Excel draws the +/- controls from `collapsed` on the summary row/column and
    sizes the outline gutter from `outlineLevelRow`/`outlineLevelCol`. Without
    both, a sheet that opens collapsed shows four rows, a handful of columns,
    and no way to expand any of it -- which is how this was first reported from
    a real Excel.

    Asserted against a saved-and-reloaded workbook, not the in-memory sheet,
    because `outlineLevelCol` only materialises on save: openpyxl's writer
    overwrites it from `column_dimensions.max_outline`. An in-memory assertion
    passes while the shipped file carries nothing.
    """
    scope = load_fixture_scope(tmp_path, _config())
    wb = Workbook()
    wb.remove(wb.active)
    build_calendar(wb, scope, _config())
    saved = tmp_path / "outline.xlsx"
    wb.save(saved)
    ws = load_workbook(saved)["Calendar"]

    assert ws.sheet_format.outlineLevelRow == 2
    assert ws.sheet_format.outlineLevelCol == 2

    by_level = {}
    for col in range(FIRST_GRID_COL, ws.max_column + 1):
        letter = ws.cell(row=1, column=col).column_letter
        by_level.setdefault(ws.column_dimensions[letter].outline_level, []).append(letter)
    assert by_level.get(0) and by_level.get(1)
    for letter in by_level[0] + by_level[1]:
        assert ws.column_dimensions[letter].collapsed is True, f"column {letter}"

    labels = _labels(ws)
    for header in (label for label in labels if str(label).startswith("BY ")):
        assert ws.row_dimensions[labels[header]].collapsed is True, header


# --- the GEB block -----------------------------------------------------------

def _geb_sheet(sources):
    """A calendar built from nothing but start dates and the GEB source field."""
    import pandas as pd

    from pipeline.report.data import build_scope
    from pipeline.scripts.process_cplan import ActivityLoad

    frame = pd.DataFrame({
        "start_date": pd.to_datetime(["2025-03-05"] * len(sources)),
        "bod_geb": list(sources),
        "activity_name": [f"Activity {i}" for i in range(len(sources))],
    })
    config = _config()
    scope = build_scope(ActivityLoad(frame, {}, {}), config)
    wb = Workbook()
    wb.remove(wb.active)
    build_calendar(wb, scope, config)
    return wb["Calendar"]


def test_the_calendar_carries_a_geb_block(tmp_path):
    ws, _ = _sheet(tmp_path)

    assert "BY GEB — multiple values possible" in _labels(ws)


def test_each_member_gets_a_row_under_the_geb_block():
    ws = _geb_sheet(["A. Person", "A. Person", "B. Person", ""])
    labels = _labels(ws)
    header = labels["BY GEB — multiple values possible"]

    assert labels["A. Person"] > header
    assert labels["B. Person"] > header
    # The unnamed activity is not dropped -- it lands in the catch-all bucket.
    assert "Not specified" in labels


def _week_total(ws, row):
    """Sum a row's literal week cells.

    The Total column is a formula by design (see the module docstring), so a
    test cannot read a number out of it without an Excel to evaluate it. The
    week cells are always literal counts, in every row -- they are what the
    formula adds up.
    """
    return sum(cell.value for cell in ws[row][FIRST_GRID_COL - 1:]
               if isinstance(cell.value, int))


def test_a_member_row_totals_only_that_members_activities():
    ws = _geb_sheet(["A. Person", "A. Person", "B. Person"])
    labels = _labels(ws)

    assert _week_total(ws, labels["A. Person"]) == 2
    assert _week_total(ws, labels["B. Person"]) == 1


def test_the_geb_block_header_counts_activities_once_despite_the_overlap():
    """Two members on one activity is two member rows but one activity."""
    ws = _geb_sheet(["A. Person; B. Person"])
    labels = _labels(ws)

    assert _week_total(ws, labels["A. Person"]) == 1
    assert _week_total(ws, labels["B. Person"]) == 1
    assert _week_total(ws, labels["BY GEB — multiple values possible"]) == 1


def test_a_detail_row_names_the_members_behind_the_activity():
    ws = _geb_sheet(["A. Person; B. Person"])
    labels = [ws.cell(row=r, column=LABEL_COL).value or ""
              for r in range(3, ws.max_row + 1)]

    assert any(label.strip() == "Activity 0 — A. Person; B. Person" for label in labels)


def test_a_detail_row_without_members_keeps_the_bare_activity_name():
    ws = _geb_sheet([""])
    detail = [ws.cell(row=r, column=LABEL_COL).value or ""
              for r in range(3, ws.max_row + 1)
              if str(ws.cell(row=r, column=LABEL_COL).value or "").startswith("  Activity")]

    assert detail
    # The block titles carry an em dash of their own, so only the detail rows
    # can answer whether an empty field left a dangling separator behind.
    assert all(label.strip() == "Activity 0" for label in detail)


# --- the audience block replaces reach --------------------------------------

def test_the_calendar_leads_with_an_audience_block_not_a_reach_one(tmp_path):
    ws, _ = _sheet(tmp_path)
    labels = _labels(ws)

    assert "BY AUDIENCE" in labels
    for gone in ("BY REACH", "Group-wide", "Multi-division", "Single division",
                 "Regional only", "Unclassified"):
        assert gone not in labels


def test_every_band_present_in_the_scope_gets_a_row(tmp_path):
    ws, scope = _sheet(tmp_path)
    labels = _labels(ws)
    header = labels["BY AUDIENCE"]

    for band in set(scope.frame["audience_band"]):
        assert band in labels, f"no row for {band!r}"
        assert labels[band] > header


def test_the_bands_are_listed_smallest_first_with_unknown_last(tmp_path):
    from pipeline.report.calendar_sheet import AUDIENCE_BAND_ORDER

    ws, scope = _sheet(tmp_path)
    labels = _labels(ws)
    present = [band for band in AUDIENCE_BAND_ORDER if band in labels]

    assert present == sorted(present, key=lambda b: labels[b])
    assert len(present) > 1, "the fixture must carry more than one band"


def test_the_audience_block_is_a_partition_so_its_members_sum_to_the_scope(tmp_path):
    """The whole reason this block keeps a vertical SUM: every activity lands
    in exactly one band, Unknown included, so the member rows really do add
    back up to the portfolio.
    """
    ws, scope = _sheet(tmp_path)
    labels = _labels(ws)
    from pipeline.report.calendar_sheet import AUDIENCE_BAND_ORDER

    total = sum(_week_total(ws, labels[band])
                for band in AUDIENCE_BAND_ORDER if band in labels)

    assert total == len(scope.frame)
