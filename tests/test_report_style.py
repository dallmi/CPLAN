"""The shared write primitives every sheet is composed from."""

import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook

from pipeline.report import style


def _sheet():
    wb = Workbook()
    return wb.active


def test_write_header_row_returns_the_next_row_and_styles_the_cells():
    ws = _sheet()

    nxt = style.write_header_row(ws, 1, ["A", "B"])

    assert nxt == 2
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=1, column=1).font.bold is True


def test_write_data_rows_applies_the_format_map_and_stripes():
    ws = _sheet()

    nxt = style.write_data_rows(ws, 1, [["x", 1], ["y", 2]], fmt_map={2: style.NUM_FMT_PCT})

    assert nxt == 3
    assert ws.cell(row=1, column=2).number_format == style.NUM_FMT_PCT
    assert ws.cell(row=2, column=1).fill.start_color.rgb.endswith(style.ROW_ALT)


def test_write_data_rows_falls_back_to_a_number_format_by_type():
    ws = _sheet()

    style.write_data_rows(ws, 1, [["x", 3, 1.5]])

    assert ws.cell(row=1, column=2).number_format == style.NUM_FMT_INT
    assert ws.cell(row=1, column=3).number_format == style.NUM_FMT_RATIO


def test_write_section_header_spans_the_requested_columns():
    ws = _sheet()

    nxt = style.write_section_header(ws, 1, "VOLUME", 3)

    assert nxt == 2
    assert ws.cell(row=1, column=1).value == "VOLUME"
    assert ws.cell(row=1, column=3).fill.start_color.rgb.endswith(style.PASTEL_I)


def test_write_formula_writes_the_formula_verbatim():
    ws = _sheet()

    style.write_formula(ws, 2, 2, "=IF(B1=0,0,B2/B1)", fmt=style.NUM_FMT_PCT)

    assert ws.cell(row=2, column=2).value == "=IF(B1=0,0,B2/B1)"
    assert ws.cell(row=2, column=2).number_format == style.NUM_FMT_PCT


def test_note_missing_writes_one_explanatory_cell():
    ws = _sheet()

    style.note_missing(ws, "No audience data available (audience column missing)")

    assert ws.cell(row=1, column=1).value.startswith("No audience data")


def test_finalize_sheet_freezes_and_clamps_column_widths():
    ws = _sheet()
    ws.cell(row=1, column=1, value="x")
    ws.cell(row=2, column=1, value="y" * 200)

    style.finalize_sheet(ws, freeze="A2")

    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["A"].width == style.MAX_WIDTH


def test_write_data_rows_does_not_number_format_booleans():
    ws = _sheet()

    style.write_data_rows(ws, 1, [["x", True, 3]])

    # bool is checked before int, so True must not pick up the integer format
    assert ws.cell(row=1, column=2).number_format != style.NUM_FMT_INT
    assert ws.cell(row=1, column=3).number_format == style.NUM_FMT_INT


def test_explicit_widths_win_over_the_auto_fit():
    ws = _sheet()
    ws.cell(row=1, column=1, value="y" * 200)

    style.finalize_sheet(ws, freeze="A2", widths={"A": 15})

    assert ws.column_dimensions["A"].width == 15


def test_write_kpi_row_accepts_a_none_value_for_a_label_only_row():
    ws = _sheet()

    nxt = style.write_kpi_row(ws, 1, "Section label", None)

    assert nxt == 2
    assert ws.cell(row=1, column=1).value == "Section label"
    assert ws.cell(row=1, column=2).value is None
